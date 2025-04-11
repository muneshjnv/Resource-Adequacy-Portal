from base64 import decode
from datetime import timedelta, datetime
from functools import wraps
from operator import itemgetter
import sys
import re

from flask import Flask, request, send_file
from flask import jsonify
from flask_cors import CORS
# import ldap
# import requests

from flask_jwt_extended import create_access_token, get_jwt
from flask_jwt_extended import create_refresh_token
from flask_jwt_extended import get_jwt_identity
from flask_jwt_extended import jwt_required
from flask_jwt_extended import JWTManager
from flask_jwt_extended import decode_token
import openpyxl
import requests
from configBackend import *
# from configBackend import generate_time_slots

import numpy as np
import pandas as pd
import json
import jwt
from flask_cors import CORS
import psycopg2
from calendar import monthrange



import datetime
from datetime import timedelta, datetime
import warnings
warnings.filterwarnings('ignore')

# APIs

from flask import Flask, request, jsonify, redirect, make_response
from werkzeug.utils import secure_filename
import os

import secrets


conn.autocommit = True
cursor = conn.cursor()


# cur = conn2.cursor()

# @app.before_request
# def force_https():
#     if not request.is_secure:
#         return redirect(request.url.replace('http://', 'https://'))


def validate_session(session_token, ip_address, device_id):
    cursor.execute("""
        SELECT expires_at, ip_address, device_id
        FROM user_sessions
        WHERE session_token = %s
    """, (session_token,))
    session_data = cursor.fetchone()

    if not session_data:
        raise Exception("Invalid session token")

    expires_at, stored_ip, stored_device = session_data
    current_time = datetime.utcnow()

    if current_time > expires_at:
        raise Exception("Session token expired")

    if stored_ip != ip_address or stored_device != device_id:
        raise Exception("Session hijacking detected")

    return True

def validate_captcha(captcha_token):
    """Validate CAPTCHA token with the CAPTCHA provider."""
    secret_key = app.config.get('RECAPTCHA_SECRET_KEY')  # Your CAPTCHA secret key
    verify_url = "https://www.google.com/recaptcha/api/siteverify"

    payload = {
        "secret": secret_key,
        "response": captcha_token
    }

    try:
        response = requests.post(verify_url, data=payload)
        result = response.json()
        return result.get("success", False)  # CAPTCHA is valid if "success" is True
    except Exception as e:
        log_error("captcha_validation", e)
        return False

# Rate limiting parameters
RATE_LIMIT = 5  # Max login attempts
RATE_LIMIT_WINDOW = timedelta(minutes=5)  # Time window for rate limiting

def is_rate_limited(ip_address, username, device_id):
    """Check if the IP address, username, and device ID are rate-limited."""
    current_time = datetime.utcnow()
    window_start = current_time - RATE_LIMIT_WINDOW

    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Count login attempts for the device ID within the time window
        cursor.execute(
            """
            SELECT COUNT(*) 
            FROM login_attempts 
            WHERE device_id = %s AND username = %s AND attempt_time > %s
            """,
            (device_id, username, window_start)
        )
        attempt_count = cursor.fetchone()[0]

        # Check if the rate limit is exceeded
        if attempt_count >= RATE_LIMIT:
            return True

        # Log the current login attempt
        cursor.execute(
            """
            INSERT INTO login_attempts (ip_address, username, device_id, attempt_time) 
            VALUES (%s, %s, %s, %s)
            """,
            (ip_address, username, device_id, current_time)
        )
        conn.commit()
    finally:
        cursor.close()
        conn.close()

    return False

# app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def log_error(api_name, error):
    exc_type, exc_obj, exc_tb = sys.exc_info()
    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
    line_number = exc_tb.tb_lineno
    logger.error("Error in %s: %s at %s:%d", api_name, error, fname, line_number)


def subtract_months(source_date, months):
    month = source_date.month - 1 - months
    year = source_date.year + month // 12
    month = month % 12 + 1
    day = min(source_date.day, [31, 29 if year % 4 == 0 and not year % 400 == 0 else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month-1])
    return datetime(year, month, day)

# Function to generate the date range with each date repeating 96 times
def generate_date_range(from_date, to_date):
    num_days = (to_date - from_date).days + 1
    date_list = []
    for single_date in (from_date + timedelta(n) for n in range(num_days)):
        date_list.extend([single_date] * 96)  # Repeat each date 96 times
    return date_list


# Helper function to generate empty columns
def _column(df, col_name):
    df[col_name] = '--'
    return df

# Function to calculate column statistics
def calculate_column_stats(column):
    pos_sum = ( column[column > 0].sum() ) /1000
    neg_sum = ( column[column < 0].sum() ) /1000
    pos_max = ( column[column > 0].max() ) *4 if len(column[column > 0]) > 0 else 0
    neg_max = ( column[column < 0].min() )*4  if len(column[column < 0]) > 0 else 0

    return pd.Series([pos_sum, neg_sum, pos_max, neg_max], index=['Pos_Sum', 'Neg_Sum', 'Pos_Max', 'Neg_Max'])

def split_string(string):
    return re.findall(r'\(([A-Za-z0-9-]+)\)', string)

def generate_dataframe_forrange(start_date,end_date):
      data = []
      current_date = start_date
      # while current_date <= end_date:
      while current_date<=end_date:
            current_time = datetime.combine(current_date, datetime.min.time())
            end_time = datetime.combine(current_date, datetime.min.time()) + timedelta(days=1)
            while current_time < end_time:
                  data.append({
                  'DATE': current_date,
                  'TIME': current_time.time()
                  })
                  current_time += timedelta(minutes=15)
            
            current_date += timedelta(days=1)
      
      df = pd.DataFrame(data)
      return df


# Function to check start and end date for column
def check_start_enddate_col(full_data_df , start_date_time , end_date_time,col_name):
      try:
            full_data_df['DATETIME'] = pd.to_datetime(full_data_df['DATE'].astype(str) + ' ' + full_data_df['TIME'].astype(str))
                                               
            temp_endtimestamp=pd.to_datetime(end_date_time,errors='coerce')
            
            if pd.isnull(temp_endtimestamp):
                  default_endtimestamp = pd.to_datetime('2050-01-01')
            else: default_endtimestamp = temp_endtimestamp

            record_dt_time_condition=full_data_df['DATETIME'].apply(lambda dt: dt >= start_date_time and dt <= default_endtimestamp )
            
            full_data_df.loc[~(record_dt_time_condition), col_name] = '--'
            full_data_df.drop(columns=['DATETIME'],inplace=True)

            return full_data_df
      except Exception as e:
            return full_data_df


def calculate_formula(row,short_locations,formula):
    try:
        for loc in short_locations:
                formula=formula.replace(loc , str(row[loc]))
        return eval(formula)
    except Exception as e:
        # extractdb_errormsg(e)
        return '--'
    


from functools import wraps
from flask import request, jsonify
from datetime import datetime
import psycopg2

def session_token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        try:
            # Get session token from the request headers
            session_token = request.headers.get('X-Session-Token')
            # print(f"Session Token Received: {session_token}")  # Debugging

            if not session_token:
                # print("Session token is missing")  # Debugging
                return jsonify({'message': 'Session token is missing'}), 401

            # Get client IP and device ID for validation
            print(request.headers)
            ip_address = request.remote_addr
            device_id = request.headers.get('Device-ID', 'unknown_device')
            # print(f"Client IP: {ip_address}, Device ID: {device_id}")  # Debugging

            # Database connection
            conn = psycopg2.connect(**db_params)
            cursor = conn.cursor()

            # Query session token details
            cursor.execute("""
                SELECT expires_at, ip_address, device_id, is_active
                FROM user_sessions
                WHERE session_token = %s
            """, (session_token,))
            session_data = cursor.fetchone()
            # print(f"Session Data from DB: {session_data}")  # Debugging

            if not session_data:
                print("Invalid session token")  # Debugging
                return jsonify({'message': 'Invalid session token'}), 401

            expires_at, stored_ip, stored_device, is_active = session_data
            current_time = datetime.utcnow()
            # print(f"Current Time: {current_time}, Expires At: {expires_at}")  # Debugging
            # print(f"Stored IP: {stored_ip}, Stored Device: {stored_device}, Is Active: {is_active}")  # Debugging

            # Check if session is active
            if not is_active:
                # print("Session has been invalidated")  # Debugging
                return jsonify({'message': 'Session has been invalidated'}), 401

            # Check if session has expired
            if current_time > expires_at:
                # print("Session token has expired")  # Debugging
                return jsonify({'message': 'Session token has expired'}), 401

            # Validate IP address and device ID (if required)
            if stored_ip != ip_address or stored_device != device_id:
                # print(f"IP/Device mismatch: Request IP={ip_address}, DB IP={stored_ip}, "
                    #   f"Request Device={device_id}, DB Device={stored_device}")  # Debugging
                return jsonify({'message': 'Session validation failed: mismatched IP or device'}), 401

            # print("Session validation passed")  # Debugging
            return f(*args, **kwargs)

        except Exception as e:
            # Log the error and return a generic error message
            # print(f"Session Validation Error: {str(e)}")  # Debugging
            log_error("session_token_validation", e)
            return jsonify({'message': 'Session validation failed'}), 500

    return decorated


# Custom token_required decorator for script access
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        jwt_data = get_jwt()

        # Check for the 'script_access' claim in the token
        # print(jwt_data, "This is JWT Data")
        if jwt_data.get('script_access'):
            # Only allow access to the /data endpoint
            if request.endpoint not in ['get_data', 'get_forecast_data']:  # 'get_data' is the function name for /data route
                return jsonify({'message': 'Access denied: this token is not authorized for this endpoint'}), 403

        return f(*args, **kwargs)

    return decorated


@app.route("/api/login", methods=["POST"])
def login():
    try:
        # Get client IP address
        ip_address = request.remote_addr

        # Get device ID from headers
        device_id = request.headers.get('Device-ID', 'unknown_device')

        # Get username and password from request
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        captcha_token = data.get('recaptcha')

        # Validate CAPTCHA
        if not validate_captcha(captcha_token):
            return jsonify({
                "error": "Invalid CAPTCHA. Please try again.",
                "status": "failure"
            })

        # Rate limiting check
        if is_rate_limited(ip_address, username, device_id):
            return jsonify({
                "error": "Too many login attempts. Please try again in 5 minutes.",
                "retry_after": RATE_LIMIT_WINDOW.seconds // 60,
                "status": "failure"
            })

        if not username or not password:
            return jsonify({"error": "Username and password are required.", "status": "failure"})

        # Database connection
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Query user data
        query = "SELECT username, password_hash, user_role, state_name, state_id FROM states WHERE username = %s"
        cursor.execute(query, (username,))
        login_data = cursor.fetchone()

        if not login_data:
            return jsonify({"error": "Invalid Credentials!", "status": "failure"})

        # Extract user details
        db_username, enc_password, role, state_name, state_id = login_data

        # Verify password
        if not bcrypt.check_password_hash(enc_password, password):
            return jsonify({
                "error": "Entered Password is Incorrect, Please try again!",
                "status": "failure"
            })

        # Generate Session Token
        session_token = secrets.token_urlsafe(32)  # Secure random session token
        session_expiration = datetime.utcnow() + timedelta(minutes=30)  # Session valid for 30 minutes

        # Store session details in the database
        cursor.execute("""
            INSERT INTO user_sessions (username, session_token, ip_address, device_id, user_agent, created_at, expires_at)
            VALUES (%s, %s, %s, %s, %s, NOW(), %s)
        """, (
            username, session_token, ip_address, device_id, request.headers.get('User-Agent'), session_expiration
        ))
        conn.commit()

        # Generate JWT Token
        jwt_data = {
            "sub": username,
            "exp": datetime.utcnow() + app.config.get('JWT_ACCESS_TOKEN_EXPIRES'),
            "role": role
        }

        if username == 'sr_internal' and password == 'Srldc#$1234':  # Special credentials
            jwt_data["script_access"] = True

        access_token = jwt.encode(payload=jwt_data, key=app.config.get('JWT_SECRET_KEY'), algorithm=app.config.get('ALGORITHM'))

        return jsonify({
            "token": access_token,            # JWT for API access
            "session_token": session_token,  # Session token for session handling
            "user": username,
            "username": state_name,
            "role": role,
            "state_id": state_id,
            "expires_at": session_expiration.strftime("%Y-%m-%d %H:%M:%S"),
            "status": "success"
        })

    except Exception as e:
        # Log the error (custom log_error function)
        log_error("login", e)
        return jsonify({"error": "Problem in logging in, Please contact ERLDC IT!", "status": "failure"})





@app.route('/api/fetchrevisions', methods=['POST'])
@jwt_required()
@token_required
def fetchRevisions():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        date = params["date"]
        params["date"] = date
        state = params["state"]
        cursor.execute("select revision_no from file_uploads where state_id = {0} and upload_date = to_date('{1}', 'DD/MM/YYYY')".format(params["state"], params["date"]))
        revisions_data = cursor.fetchall()

        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for {0}".format(date), revisions=revision_list)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for {0} state for date {1}".format(state_name, date))
    except Exception as e:
        log_error("fetchrevisions", e)
        return jsonify(message="There is Some Problem, Please contact ERLDC IT")


@app.route('/api/fetchintradayrevisions', methods=['POST'])
@jwt_required()
@token_required
def fetchIntradayRevisions():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        date = params["date"]
        params["date"] = date
        state = params["state"]
        cursor.execute("select revision_no from intraday_file_uploads where state_id = {0} and upload_date = to_date('{1}', 'DD/MM/YYYY')".format(params["state"], params["date"]))
        revisions_data = cursor.fetchall()

        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for {0}".format(date), revisions=revision_list)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for {0} state for date {1}".format(state_name, date))
    except Exception as e:
        log_error("fetchrevisions", e)
        return jsonify(message="There is Some Problem, Please contact ERLDC IT")


@app.route('/api/fetchweekrevisions', methods=['POST'])
@jwt_required()
@token_required
def fetchWeekRevisions():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        from_date = params["from_date"]
        to_date = params["to_date"]
        state = params["state"]
        cursor.execute("select revision_no from week_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY')".format(params["state"], from_date, to_date))
        revisions_data = cursor.fetchall()
        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for '{0}'-'{1}'".format(from_date, to_date), revisions=revision_list, from_date=from_date, to_date=to_date)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for state {0} for week '{1}'-'{2}'".format(state_name, from_date, to_date))    
    
    except Exception as e:
        log_error("fetchweekrevisions", e)
        return jsonify(message="There is Some Problem, Please contact ERLDC IT")

@app.route('/api/fetchweeklyrevisionsdata', methods=['POST'])
@jwt_required()
@token_required
def fetchWeeklyRevisionsData():
    try:

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        print(params)
        cursor.execute("select file_data, from_date, to_date, upload_time, uploaded_by, revision_no from week_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY') and revision_no={3}".format(params["state"], params["from_date"], params["to_date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        from_date = ''
        to_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            from_date = data[0][1].strftime('%d-%b-%Y')
            to_date = data[0][2].strftime('%d-%b-%Y')
            uploaded_time = data[0][3].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][4]
            revision_no = data[0][5]

            # print(len(file_data), len(file_data[0]))
            cursor.close()
            return jsonify(status="success", data=file_data, time=uploaded_time, from_date=from_date,to_date=to_date, revision=revision_no, role=uploaded_by)
        else:
            cursor.close()
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact ERLDC IT!")
            

        # return jsonify(message="Fetched Successfully")

    except Exception as e:
        log_error("weeklyrevisionsdata", e)
        return jsonify(message="There is Some Problem, Please contact ERLDC IT")
    

@app.route('/api/fetchintradayrevisionsdata', methods=['POST'])
@jwt_required()
@token_required
def checkIntradayUploaded():
    try:

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)
        cursor.execute("select file_data, upload_date, upload_time, uploaded_by, revision_no from intraday_file_uploads where state_id = {0} and upload_date = to_date('{1}', 'DD/MM/YYYY') and revision_no={2}".format(params["state"], params["date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        uploaded_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            uploaded_date = data[0][1].strftime('%d-%b-%Y')
            uploaded_time = data[0][2].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][3]
            revision_no = data[0][4]
            # print(type(uploaded_date), type(uploaded_time))
            # print(type(file_data), uploaded_time, uploaded_date, uploaded_by, revision_no)
            # print(file_data[-1])
            # print(len(file_data), len(file_data[0]))
            cursor.close()
            return jsonify(status="success", data=file_data, time=uploaded_time, date=uploaded_date, revision=revision_no, role=uploaded_by)
        else:
            cursor.close()
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact ERLDC IT!")
    except Exception as e:
        log_error("fetchintradayrevisionsdata", e)
        
 


@app.route('/api/fetchrevisionsdata', methods=['POST'])
@jwt_required()
@token_required
def checkUploaded():
    try:

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)
        cursor.execute("select file_data, upload_date, upload_time, uploaded_by, revision_no from file_uploads where state_id = {0} and upload_date = to_date('{1}', 'DD/MM/YYYY') and revision_no={2}".format(params["state"], params["date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        uploaded_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            uploaded_date = data[0][1].strftime('%d-%b-%Y')
            uploaded_time = data[0][2].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][3]
            revision_no = data[0][4]
            # print(type(uploaded_date), type(uploaded_time))
            # print(type(file_data), uploaded_time, uploaded_date, uploaded_by, revision_no)
            # print(file_data[-1])
            # print(len(file_data), len(file_data[0]))
            cursor.close()
            return jsonify(status="success", data=file_data, time=uploaded_time, date=uploaded_date, revision=revision_no, role=uploaded_by)
        else:
            cursor.close()
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact ERLDC IT!")
    except Exception as e:
        log_error("fetchrevisionsdata", e)
        
        

    # return jsonify(message="Fetched Successfully")


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


state_revision_numbers = {}


@app.route('/api/uploadintraday', methods=['POST'])
@jwt_required()
@token_required
def uploadIntradayDataAndFile():
    try:
        # Connect to the database
        conn = psycopg2.connect(
            database="demand_forecast_states", user='ra_admin', 
            password='admin', host='localhost', port='5432'
        )
        cursor = conn.cursor()

        # Fetch header and form data
        header_data = dict(request.headers)
        state = request.form.get('state')
        disabledDate = request.form.get('disabledDate')
        data = json.loads(request.form.get('data'))
        data = json.dumps(data)

        # Decode token to fetch role
        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)
        role = x['role']

        # Fetch the state name
        cursor.execute("SELECT state_name FROM states WHERE state_id = %s", (state,))
        state_name = cursor.fetchone()[0]

        # Parse the date from the format received
        date_format = "%a %b %d %Y %H:%M:%S GMT%z (%Z)"
        disabledDate = datetime.strptime(disabledDate, date_format).strftime("%Y-%m-%d")

        # Ensure file is present in the request
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})

        file = request.files['excelFile']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        if file and allowed_file(file.filename):
            # Save the uploaded file
            file_name = secure_filename(file.filename)
            directory_path = os.path.join(shared_drive_path, "INTRADAY", disabledDate, state_name)
            os.makedirs(directory_path, exist_ok=True)

            # Check if this is the first revision
            cursor.execute(
                "SELECT * FROM intraday_file_uploads WHERE upload_date = to_date(%s, 'YYYY-MM-DD') AND state_id = %s", 
                (disabledDate, state)
            )
            existing_revs = cursor.fetchall()
            revision_no = len(existing_revs)

            # Save the file with the appropriate revision number
            filename = f"{disabledDate}_{state_name}_rev{revision_no}.xlsx"
            file_path = os.path.join(directory_path, filename)
            file.save(file_path)

            # Insert the file data into the database
            cursor.execute(
                """
                INSERT INTO intraday_file_uploads (state_id, upload_date, upload_time, file_name, revision_no, uploaded_by, file_data) 
                VALUES (%s, to_date(%s, 'YYYY-MM-DD'), to_timestamp(%s, 'YYYY-MM-DD HH24:MI:SS'), %s, %s, %s, %s)
                """, 
                (state, disabledDate, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, revision_no, role, data)
            )
            conn.commit()

            # Check if all states except the current one have uploaded for the given date
            cursor.execute(
                """
                SELECT state_id FROM states WHERE state_id NOT IN (
                    SELECT state_id FROM intraday_file_uploads WHERE upload_date = %s
                )
                AND state_id IN (1,2,3,4,5,7)  -- Exclude some states as required
                """, 
                (disabledDate,)
            )
            pending_states = cursor.fetchall()

            # If no pending states are left and this is the first upload for the current state, trigger the email
            if not pending_states and revision_no == 0 and False:
                # Fetch data for the consolidated file (using the updated function)
                db_data = fetch_max_revision_data(for_week_ahead=False, custom_date=disabledDate)

                if db_data:
                    # file_path_template = r'D:\Demand Forecasting Page\custom scripts\day_ahead_consolidated.xlsx'
                    file_path_template = os.path.join(shared_drive_path, "FORMATS", "day_ahead_consolidated.xlsx")
                    save_file_path = write_to_excel(file_path_template, db_data, for_week_ahead='day', custom_date=disabledDate)
                    
                    # Send the email with the attached consolidated file
                    try:
                        send_mails_list_db = ['mprashada@grid_india.in']
                        cc_mails_list_db = ['srldcosa@grid_india.in']
                        send_mail_with_attachment(send_mails_list_db, cc_mails_list_db, save_file_path, for_week_ahead='day', custom_date=disabledDate)
                        print(f"All states have uploaded. Email with consolidated file sent.")
                    except Exception as email_error:
                        # Log any failure during email sending, but continue with the upload success response
                        print(f"Failed to send email: {email_error}")
                        log_error("Email sending failed", email_error)

            # Return success response for the file upload
            return jsonify({'message': f'Data and file uploaded successfully. Uploaded for Revision-{revision_no}'})

        else:
            return jsonify({'error': 'Invalid file type'})

    except Exception as e:
        log_error("uploadintraday", e)
        return jsonify(message="There is a problem in uploading. Please contact ERLDC IT!")

    finally:
        # Ensure that resources are closed properly
        cursor.close()
        conn.close()



@app.route('/api/uploaddayahead', methods=['POST'])
@jwt_required()
@token_required
def uploadDayAheadDataAndFile():
    try:
        # Connect to the database
        conn = psycopg2.connect(
            database="demand_forecast_states", user='ra_admin', 
            password='admin', host='localhost', port='5432'
        )
        cursor = conn.cursor()

        # Fetch header and form data
        header_data = dict(request.headers)
        state = request.form.get('state')
        disabledDate = request.form.get('disabledDate')
        data = json.loads(request.form.get('data'))
        data = json.dumps(data)

        # Decode token to fetch role
        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)
        role = x['role']

        # Fetch the state name
        cursor.execute("SELECT state_name FROM states WHERE state_id = %s", (state,))
        state_name = cursor.fetchone()[0]

        # Parse the date from the format received
        date_format = "%a %b %d %Y %H:%M:%S GMT%z (%Z)"
        disabledDate = datetime.strptime(disabledDate, date_format).strftime("%Y-%m-%d")

        # Ensure file is present in the request
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})

        file = request.files['excelFile']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        if file and allowed_file(file.filename):
            # Save the uploaded file
            file_name = secure_filename(file.filename)
            directory_path = os.path.join(shared_drive_path, "DAY_AHEAD", disabledDate, state_name)
            os.makedirs(directory_path, exist_ok=True)

            # Check if this is the first revision
            cursor.execute(
                "SELECT * FROM file_uploads WHERE upload_date = to_date(%s, 'YYYY-MM-DD') AND state_id = %s", 
                (disabledDate, state)
            )
            existing_revs = cursor.fetchall()
            revision_no = len(existing_revs)

            # Save the file with the appropriate revision number
            filename = f"{disabledDate}_{state_name}_rev{revision_no}.xlsx"
            file_path = os.path.join(directory_path, filename)
            file.save(file_path)

            # Insert the file data into the database
            cursor.execute(
                """
                INSERT INTO file_uploads (state_id, upload_date, upload_time, file_name, revision_no, uploaded_by, file_data) 
                VALUES (%s, to_date(%s, 'YYYY-MM-DD'), to_timestamp(%s, 'YYYY-MM-DD HH24:MI:SS'), %s, %s, %s, %s)
                """, 
                (state, disabledDate, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, revision_no, role, data)
            )
            conn.commit()

            # Check if all states except the current one have uploaded for the given date
            cursor.execute(
                """
                SELECT state_id FROM states WHERE state_id NOT IN (
                    SELECT state_id FROM file_uploads WHERE upload_date = %s
                )
                AND state_id IN (1,2,3,4,5,7)  -- Exclude some states as required
                """, 
                (disabledDate,)
            )
            pending_states = cursor.fetchall()

            # If no pending states are left and this is the first upload for the current state, trigger the email
            if not pending_states and revision_no == 0:
                # Fetch data for the consolidated file (using the updated function)
                db_data = fetch_max_revision_data(for_week_ahead=False, custom_date=disabledDate)

                if db_data:
                    # file_path_template = r'D:\Demand Forecasting Page\custom scripts\day_ahead_consolidated.xlsx'
                    file_path_template = os.path.join(shared_drive_path, "FORMATS", "day_ahead_consolidated.xlsx")
                    save_file_path = write_to_excel(file_path_template, db_data, for_week_ahead='day', custom_date=disabledDate)
                    
                    # Send the email with the attached consolidated file
                    try:
                        send_mails_list_db = ['mprashad@grid_india.in']
                        cc_mails_list_db = ['srldcos@grid_india.in']
                        send_mail_with_attachment(send_mails_list_db, cc_mails_list_db, save_file_path, for_week_ahead='day', custom_date=disabledDate)
                        print(f"All states have uploaded. Email with consolidated file sent.")
                    except Exception as email_error:
                        # Log any failure during email sending, but continue with the upload success response
                        print(f"Failed to send email: {email_error}")
                        log_error("Email sending failed", email_error)

            # Return success response for the file upload
            return jsonify({'message': f'Data and file uploaded successfully. Uploaded for Revision-{revision_no}'})

        else:
            return jsonify({'error': 'Invalid file type'})

    except Exception as e:
        log_error("uploaddayahead", e)
        return jsonify(message="There is a problem in uploading. Please contact ERLDC IT!")

    finally:
        # Ensure that resources are closed properly
        cursor.close()
        conn.close()


        
        # return jsonify(message="There is problem in uploading, Please contact ERLDC IT!")




@app.route("/api/downloadintraday", methods=["POST"])
@jwt_required()
@token_required
def downloadIntraday():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Parse the JSON input from the frontend
        data_json = request.get_json()
        state_id = data_json.get('state')
        # upload_date = data_json.get('date')  # Expected in 'YYYY-MM-DD' format
        upload_date_str = data_json.get('date')  # Expected in 'YYYY-MM-DD' format
        upload_date = datetime.strptime(upload_date_str, "%d/%m/%Y").date()
        revision_no = data_json.get('revision')

        print(data_json)

        # Query the database to fetch the file_name (which is the file path)
        query = """
            SELECT file_name 
            FROM public.intraday_file_uploads 
            WHERE state_id = %s 
            AND upload_date = %s 
            AND revision_no = %s
        """
        cursor = conn.cursor()
        cursor.execute(query, (state_id, upload_date, revision_no))
        result = cursor.fetchone()

        if not result:
            print("File not found")
            return jsonify({'status': 'failure', 'error': 'File not found'}), 404

        # Extract the file path
        file_path = result[0]
        print(file_path)
        # Ensure the file exists before sending it
        if not os.path.exists(file_path):
            print("Path does not exists")
            cursor.close()
            return jsonify({'status': 'failure', 'error': 'File not found on server'}), 404

        # Send the file as an attachment
        return send_file(file_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        log_error("downloadintraday", e)
        cursor.close()
        
        # cursor.close()
        return jsonify({'status': 'failure', 'error': str(e)}), 500

@app.route("/api/downloaddayahead", methods=["POST"])
@jwt_required()
@token_required
def downloadDayAhead():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Parse the JSON input from the frontend
        data_json = request.get_json()
        state_id = data_json.get('state')
        upload_date_str = data_json.get('date')  # Expected in 'YYYY-MM-DD' format
        upload_date = datetime.strptime(upload_date_str, "%d/%m/%Y").date()
        revision_no = data_json.get('revision')

        print(data_json)

        # Query the database to fetch the file_name (which is the file path)
        query = """
            SELECT file_name 
            FROM public.file_uploads 
            WHERE state_id = %s 
            AND upload_date = %s 
            AND revision_no = %s
        """
        cursor = conn.cursor()
        cursor.execute(query, (state_id, upload_date, revision_no))
        result = cursor.fetchone()

        if not result:
            print("File not found")
            return jsonify({'status': 'failure', 'error': 'File not found'}), 404

        # Extract the file path
        file_path = result[0]
        print(file_path)

        # Ensure the file exists before sending it
        if not os.path.exists(file_path):
            print("Path does not exists")
            cursor.close()
            return jsonify({'status': 'failure', 'error': 'File not found on server'}), 404

        # Send the file as an attachment
        return send_file(file_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        log_error("downloaddayahead", e)
        cursor.close()
        
        # cursor.close()
        return jsonify({'status': 'failure', 'error': str(e)}), 500



@app.route("/api/downloadweekahead", methods=["POST"])
@jwt_required()
@token_required
def downloadWeekAhead():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Parse the JSON input from the frontend
        data_json = request.get_json()
        state_id = data_json.get('state')
        # upload_from_date = data_json.get('from_date')  # Expected in 'YYYY-MM-DD' format
        upload_from_date_str = data_json.get('from_date')  # Expected in 'YYYY-MM-DD' format
        upload_from_date = datetime.strptime(upload_from_date_str, "%d/%m/%Y").date()
        # upload_to_date = data_json.get('to_date')
        upload_to_date_str = data_json.get('to_date')  # Expected in 'YYYY-MM-DD' format
        upload_to_date = datetime.strptime(upload_to_date_str, "%d/%m/%Y").date()
        revision_no = data_json.get('revision')

        print(data_json)

        # Query the database to fetch the file_name (which is the file path)
        query = """
            SELECT file_name 
            FROM public.week_ahead_file_uploads 
            WHERE state_id = %s 
            AND from_date = %s
            AND to_date = %s 
            AND revision_no = %s
        """
        cursor = conn.cursor()
        cursor.execute(query, (state_id, upload_from_date, upload_to_date, revision_no))
        result = cursor.fetchone()

        if not result:
            print("File not found")
            return jsonify({'status': 'failure', 'error': 'File not found'}), 404

        # Extract the file path
        file_path = result[0]
        print(file_path)

        # Ensure the file exists before sending it
        if not os.path.exists(file_path):
            print("Path does not exists")
            cursor.close()
            return jsonify({'status': 'failure', 'error': 'File not found on server'}), 404

        # Send the file as an attachment
        return send_file(file_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        log_error("downloadweekahead", e)
        cursor.close()
        
        # cursor.close()
        return jsonify({'status': 'failure', 'error': str(e)}), 500


@app.route("/api/downloadmonthahead", methods=["POST"])
@jwt_required()
@token_required
def downloadMonthAhead():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Parse the JSON input from the frontend
        data_json = request.get_json()
        state_id = data_json.get('state')
        # upload_from_date = data_json.get('from_date')  # Expected in 'YYYY-MM-DD' format
        upload_from_date_str = data_json.get('from_date')  # Expected in 'YYYY-MM-DD' format
        upload_from_date = datetime.strptime(upload_from_date_str, "%d/%m/%Y").date()
        # upload_to_date = data_json.get('to_date')
        upload_to_date_str = data_json.get('to_date')  # Expected in 'YYYY-MM-DD' format
        upload_to_date = datetime.strptime(upload_to_date_str, "%d/%m/%Y").date()
        revision_no = data_json.get('revision')

        print(data_json)

        # Query the database to fetch the file_name (which is the file path)
        query = """
            SELECT file_name 
            FROM public.month_ahead_file_uploads 
            WHERE state_id = %s 
            AND from_date = %s
            AND to_date = %s 
            AND revision_no = %s
        """
        cursor = conn.cursor()
        cursor.execute(query, (state_id, upload_from_date, upload_to_date, revision_no))
        result = cursor.fetchone()

        if not result:
            print("File not found")
            return jsonify({'status': 'failure', 'error': 'File not found'}), 404

        # Extract the file path
        file_path = result[0]
        print(file_path)

        # Ensure the file exists before sending it
        if not os.path.exists(file_path):
            print("Path does not exists")
            cursor.close()
            return jsonify({'status': 'failure', 'error': 'File not found on server'}), 404

        # Send the file as an attachment
        return send_file(file_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        log_error("downloadmonthahead", e)
        cursor.close()
        
        # cursor.close()
        return jsonify({'status': 'failure', 'error': str(e)}), 500

@app.route("/api/downloadyearahead", methods=["POST"])
@jwt_required()
@token_required
def downloadYearAhead():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Parse the JSON input from the frontend
        data_json = request.get_json()
        state_id = data_json.get('state')
        # upload_from_date = data_json.get('from_date')  # Expected in 'YYYY-MM-DD' format
        upload_from_date_str = data_json.get('from_date')  # Expected in 'YYYY-MM-DD' format
        upload_from_date = datetime.strptime(upload_from_date_str, "%d/%m/%Y").date()
        # upload_to_date = data_json.get('to_date')
        upload_to_date_str = data_json.get('to_date')  # Expected in 'YYYY-MM-DD' format
        upload_to_date = datetime.strptime(upload_to_date_str, "%d/%m/%Y").date()
        revision_no = data_json.get('revision')

        print(data_json)

        # Query the database to fetch the file_name (which is the file path)
        query = """
            SELECT file_name 
            FROM public.year_ahead_file_uploads 
            WHERE state_id = %s 
            AND from_date = %s
            AND to_date = %s 
            AND revision_no = %s
        """
        cursor = conn.cursor()
        cursor.execute(query, (state_id, upload_from_date, upload_to_date, revision_no))
        result = cursor.fetchone()

        if not result:
            print("File not found")
            return jsonify({'status': 'failure', 'error': 'File not found'}), 404

        # Extract the file path
        file_path = result[0]
        print(file_path)

        # Ensure the file exists before sending it
        if not os.path.exists(file_path):
            print("Path does not exists")
            cursor.close()
            return jsonify({'status': 'failure', 'error': 'File not found on server'}), 404

        # Send the file as an attachment
        return send_file(file_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        log_error("downloadyearahead", e)
        cursor.close()
        
        # cursor.close()
        return jsonify({'status': 'failure', 'error': str(e)}), 500

@app.route('/api/pendingentries', methods=['GET'])
@jwt_required()
@token_required
def pendingTimingEntries():
    try:
        response = requests.get("https://oms2.srldc.in/Codebook/getDueTimingEntryData")

        data = response.json()["data"]["dueTiminingEntryList"]

        # print(data[0])

        # print(data)


        # print(data[0])

        header_data = dict(request.headers)
        # print(header_data)
        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)

        username = x['sub']   # username in database

        role = x['role']       # role in database

        states_entities_dict = {}
        states_entities_dict["t_kar_state"] = ['kptcl', 'KTL']
        states_entities_dict["t_ap_state"] = ['LKPPLSTG3', 'APTRANSCO']
        states_entities_dict["t_tn_state"] = ['TN','TANTRANSCO']
        states_entities_dict["t_ker_state"] = ['kseb', 'kseb_sd']
        states_entities_dict["t_tg_state"] = ['tgtransco']
        states_entities_dict["t_pondy_state"] = []
        states_entities_dict["pgcil_sr_1"] = ['PGCIL SR-1']
        states_entities_dict["pgcil_sr_2"] = ['PGCIL SR-2']


        admin_states_list = []

        for key, value in states_entities_dict.items():
            for v in value:
                admin_states_list.append(v)


        # PGCIL SR-2, PGCIL SR-1

        # print(states_entities_dict[username])


        entities_data = []

        # print(data)

        id = 1

        if role == 'user':
            for i in data:
                if (any(item.lower() in i["codeIssuedto"].lower().split(",") for item in states_entities_dict[username])  or i["codeRequestedby"] in states_entities_dict[username]) and (i["constituentEnteredTime"] == ""):
                    entities_data.append({"id": id, "codeIssueTime": i["codeIssuedTime"], "elementType": i["entityFeatureName"], "elementName": i["elementName"], "switching": i["end"], "nldcCode": i["nldcCode"] + "/" + i["otherRegionCode"], "srldcCode": i["codeNo"], "category": i["outageCategory"], "codeIssuedTo": i["codeIssuedto"], "codeRequestedBy": i["codeRequestedby"], "codeId": i["codeId"], "isSelected": False})
                    id = id + 1
        
        else:
            for i in data:
                if (i["constituentEnteredTime"] == ""):
                    entities_data.append({"id": id, "codeIssueTime": i["codeIssuedTime"], "elementType": i["entityFeatureName"], "elementName": i["elementName"], "switching": i["end"], "nldcCode": i["nldcCode"] + "/" + i["otherRegionCode"], "srldcCode": i["codeNo"], "category": i["outageCategory"], "codeIssuedTo": i["codeIssuedto"], "codeRequestedBy": i["codeRequestedby"],  "codeId": i["codeId"], "isSelected": False})
                    id = id + 1

        # print(type(entities_data[0]["codeIssueTime"]))
        # formatted_data_list = [{'date_key': datetime.strptime(item['date_key'], '%Y-%m-%dT%H:%M:%S.%f').strftime('%d/%m/%Y, %H:%M') for item in data_list]
                               
        # for i in range(len(entities_data)):
        #     # entities_data[i]["codeIssueTime"] = datetime.strptime(entities_data[i]["codeIssueTime"], '%Y-%m-%dT%H:%M:%S.%f').isoformat()
        #     # entities_data[i]["codeIssueTime"] = entities_data[i]["codeIssueTime"]
        #     date_obj = datetime.strptime(entities_data[i]["codeIssueTime"], '%Y-%m-%d %H:%M:%S')  # Convert to datetime object
        #     entities_data[i]["codeIssueTime"] = date_obj.strftime('%d/%m/%Y, %H:%M')  # Format as 'dd/MM/yyyy, HH:mm'
        

            # print(entities_data[i]["codeIssueTime"])
        
        # print(entities_data[0]["codeIssueTime"])


        for i in range(len(entities_data)):
            # Convert to datetime object, accounting for milliseconds if present
            date_str = entities_data[i]["codeIssueTime"]

            try:
                # Attempt to parse with milliseconds
                date_obj = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S.%f')
            except ValueError:
                # Fallback to seconds-only format if milliseconds are absent
                date_obj = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S')
            
            # Convert back to a standardized ISO format (without milliseconds)
            entities_data[i]["codeIssueTime"] = date_obj.strftime('%Y-%m-%dT%H:%M:%S')


 


        if len(entities_data) == 0:
            return jsonify(status="failure", message="Data is Empty!")
        


        return jsonify(status="success", message="Message fetched successfully!", data=entities_data )


    except Exception as e:
        # print(e)
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print(exc_type, fname, exc_tb.tb_lineno)
        log_error("pendingtimingentries", e)
        cursor.close()
        return jsonify(status="failure", message="There is a problem in fetching the data, Please contact ERLDC IT!")


@app.route('/api/getelementpreviouscodes', methods=['POST'])
@jwt_required()
@token_required
def getPreviousCodes():
    try:
        params = request.get_json()

        print(params)

        params = params["params"]
        from_date = params["from_date"]
        to_date = params["to_date"]
        feature_name = params.get("feature_name", "")  # Provide a default value if not present
        pwcEntityId = params.get("pwcEntityId", "")  # Provide a default value if not present


        


        # Create the payload to send in the POST request
        payload = json.dumps({
            "startDate": from_date,
            "endDate": to_date,
            "pwcEntityId": "",
            "elementId": "",
            "featureName": ""
        })

        headers = {
            "Content-Type": "application/json"
            # "Authorization": "Bearer your_token"  # If authentication is required
        }

        response = requests.post("https://oms2.srldc.in/Codebook/getElementPreviousCodes", data=payload, headers=headers)

        # print(response.status_code)
        # print(response.json())

        header_data = dict(request.headers)
        # print(header_data)
        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)

        username = x['sub']   # username in database

        username = x['sub']   # username in database

        role = x['role']       # role in database

        states_entities_dict = {}
        states_entities_dict["t_kar_state"] = ['kptcl', 'KTL']
        states_entities_dict["t_ap_state"] = ['LKPPLSTG3', 'APTRANSCO']
        states_entities_dict["t_tn_state"] = ['TN','TANTRANSCO']
        states_entities_dict["t_ker_state"] = []
        states_entities_dict["t_tg_state"] = ['TSTRANSCO']
        states_entities_dict["t_pondy_state"] = []
        states_entities_dict["pgcil_sr_1"] = ['PGCIL SR-1']
        states_entities_dict["pgcil_sr_2"] = ['PGCIL SR-2']


        admin_states_list = []

        for key, value in states_entities_dict.items():
            for v in value:
                admin_states_list.append(v)


        # PGCIL SR-2, PGCIL SR-1

        # print(states_entities_dict[username])



        if response.status_code == 200:
            data = response.json()["data"]["previousCodes"]
            # print(data)
            # print(data['data']["previousCodes"].keys())
                    
            entities_data = []

            # print(data)

            id = 1

            if role == 'user':
                for i in data:
                    if (i["codeIssuedto"] in states_entities_dict[username] or i["codeRequestedby"] in states_entities_dict[username]):
                        entities_data.append({"id": id, "codeIssueTime": i["codeIssuedTime"], "elementType": i["entityFeatureName"], "elementName": i["elementName"], "switching": i["end"], "nldcCode": i["nldcCode"], "srldcCode": i["codeNo"], "category": i["outageCategory"], "codeIssuedTo": i["codeIssuedto"], "codeRequestedBy": i["codeRequestedby"], "codeId": i["codeId"], "isSelected": False})
                        id = id + 1
            
            else:
                for i in data:
                    # print(i.keys())
                    # print('constituentEnteredTime' in i.keys())
                    if (i["codeIssuedto"] in admin_states_list or i["codeRequestedby"] in admin_states_list):
                        entities_data.append({"id": id, "codeIssueTime": i["codeIssuedTime"], "elementType": i["entityFeatureName"], "elementName": i["elementName"], "switching": i["end"], "nldcCode": i["nldcCode"], "srldcCode": i["codeNo"], "category": i["outageCategory"], "codeIssuedTo": i["codeIssuedto"], "codeRequestedBy": i["codeRequestedby"], "codeId": i["codeId"], "isSelected": False})
                        id = id + 1

            # print(type(entities_data[0]["codeIssueTime"]))
            # formatted_data_list = [{'date_key': datetime.strptime(item['date_key'], '%Y-%m-%dT%H:%M:%S.%f').strftime('%d/%m/%Y, %H:%M') for item in data_list]
                                
            for i in range(len(entities_data)):
                entities_data[i]["codeIssueTime"] = datetime.strptime(entities_data[i]["codeIssueTime"], '%Y-%m-%dT%H:%M:%S.%f').isoformat()
            
            # print(entities_data[0]["codeIssueTime"])


            if len(entities_data) == 0:
                return jsonify(status="failure", message="Data is Empty!")
        else:
            # print(response.status_code)
            return jsonify(status="failure", message="There is a problem in fetching data, Please contact ERLDC IT!")
        return jsonify(status="success", message="Message fetched successfully!", data=entities_data )
    except Exception as e:
        # print(e)
        log_error("getpreviouscodes", e)
        cursor.close()
        return jsonify(status="failure", message="There is a problem in fetching the data, Please contact ERLDC IT!")
 

@app.route('/api/weekaheadformat')
@jwt_required()
@token_required
def weekAheadFormat():
    try:
        # Function to find the date of the next Monday
        def next_monday(d):
            # Calculate the number of days until the next Monday
            days_til_monday = (7 - d.weekday()) % 7
            if days_til_monday == 0:
                # If today is Monday, add 7 days to get the next Monday
                days_til_monday = 7
            return d + timedelta(days=days_til_monday)

        # Function to create the 2D list
        def create_2d_list(start_date, num_days, num_blocks_per_day):
            data = []

            for day in range(num_days):
                current_date = start_date + timedelta(days=day)
                for block in range(num_blocks_per_day):
                    current_time = (datetime.min + timedelta(minutes=block * 15)).time()
                    next_time = (datetime.min + timedelta(minutes=(block + 1) * 15)).time()
                    timestamp = f"{current_time:%H:%M} - {next_time:%H:%M}"
                    row = [current_date.strftime('%Y-%m-%d'), block + 1, timestamp] + [0] * 19
                    data.append(row)

            return data

        # Find the next Monday
        today = datetime.now()
        next_monday_date = next_monday(today)

        # Create the 2D list for 7 days with 96 blocks each
        num_days = 7
        num_blocks_per_day = 96
        data = create_2d_list(next_monday_date, num_days, num_blocks_per_day)

        
        return jsonify(status="success", data=data)
    
    except Exception as e:
        log_error("weekaheadformat", e)
        cursor.close()
        return jsonify(msg="Problem in fetching the data, Please contact ERLDC IT", status="failure")


@app.route('/api/monthaheadformat')
@jwt_required()
@token_required
def monthAheadFormat():

    try:

        def create_2d_list(start_date, end_date, num_blocks_per_day):
            data = []
            
            while start_date <= end_date:
                for block in range(num_blocks_per_day):
                    current_time = (datetime.min + timedelta(minutes=block * 15)).time()
                    next_time = (datetime.min + timedelta(minutes=(block + 1) * 15)).time()
                    timestamp = f"{current_time:%H:%M} - {next_time:%H:%M}"
                    row = [start_date.strftime('%Y-%m-%d'), block + 1, timestamp] + [0] * 19
                    data.append(row)
                
                start_date += timedelta(days=1)
            
            return data

        today = datetime.now()

        # Calculate the start of the next month, handling month rollover if necessary
        next_month = (today.month % 12) + 1
        next_year = today.year + (today.month // 12)

        start_of_next_month = datetime(next_year, next_month, 1)

        # Calculate the end of the next month by getting the first day of the month after next, then subtracting one day
        following_month = (next_month % 12) + 1
        following_year = next_year + (next_month // 12)

        end_of_next_month = datetime(following_year, following_month, 1) - timedelta(days=1)


        # Create the 2D list for the entire next month with 96 blocks each day
        num_blocks_per_day = 96
        data = create_2d_list(start_of_next_month, end_of_next_month, num_blocks_per_day)

        # print(len(data))

        return jsonify(data=data, status="success")
    except Exception as e:
        log_error("monthaheadformat", e)
        cursor.close()
        return jsonify(status="failure", msg="Problem in Fetching the data, Please contact ERLDC IT!")



@app.route('/api/yearaheadformat')
@jwt_required()
@token_required
def yearheadFormat():
    try:
        def create_2d_list_for_year(start_date, end_date, num_blocks_per_day):
            data = []

            while start_date <= end_date:
                for block in range(num_blocks_per_day):
                    current_time = (datetime.min + timedelta(minutes=block * 60)).time()
                    next_time = (datetime.min + timedelta(minutes=(block + 1) * 60)).time()
                    row = [start_date.strftime('%Y-%m-%d'), block + 1] + [0] * 19
                    data.append(row)

                start_date += timedelta(days=1)

            return data

        # Get the current date
        today = datetime.now()
        
        # Define the financial year starting from the next April 1st
        if today.month >= 4:
            start_of_financial_year = datetime(today.year + 1, 4, 1)
            end_of_financial_year = datetime(today.year + 2, 3, 31)
        else:
            start_of_financial_year = datetime(today.year, 4, 1)
            end_of_financial_year = datetime(today.year + 1, 3, 31)

        # num_blocks_per_day = 24
        data_for_financial_year = create_2d_list_for_year(start_of_financial_year, end_of_financial_year, num_blocks_per_day=24)

        return jsonify(data=data_for_financial_year, status="success")
    
    except Exception as e:
        log_error("yearaheadformat", e)
        cursor.close()
        return jsonify(status="failure", msg="Problem in Fetching the data, Please contact ERLDC IT!")

    
@app.route('/api/submitentries', methods=['POST'])
@jwt_required()
@token_required
def submitTimingEntries():
    try:
        data = request.get_json()

        # data = json.loads(data)

        data = data["data"]

        import pdb
        pdb.set_trace()




        for i in data:

            if 'T' in i["codeIssueTime"]:
                i["codeIssueTime"] = datetime.fromisoformat(i["codeIssueTime"]).strftime("%Y-%m-%d %H:%M")
            else:
                i["codeIssueTime"] = datetime.strptime(i["codeIssueTime"], "%d/%m/%Y, %H:%M")

                # Format the datetime object as "yyyy-mm-dd hh:mm"
                i["codeIssueTime"] = i["codeIssueTime"].strftime("%Y-%m-%d %H:%M")




            response = requests.get("https://oms2.srldc.in/Codebook/EnterConstituentTime?codeId={0}&enteredTime={1}".format(i["codeId"], i["codeIssueTime"]))

            # print(response)

        return jsonify(status="success", message="Sent Successfully!")
    except Exception as e:
        log_error("submittimingentries", e)
        cursor.close()
        # print(error)
        return jsonify(status="failure", message="There is a problem, Please contact ERLDC IT!")
    
    

@app.route('/api/weeekaheadforecast', methods=['POST'])
@jwt_required()
def weekAheadForecast():
    try: 
        return jsonify(msg="Successful!", status="failure")
    except Exception as error:
        return jsonify(msg="There is some problem, Please contact ERLDC IT!")
    


@app.route('/api/uploadweekahead', methods=['POST'])
@jwt_required()
@token_required
def uploadWeekAheadDataAndFile():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        header_data = dict(request.headers)
        state = request.form.get('state')
        fromDate = request.form.get('fromDate')
        toDate = request.form.get('toDate')
        data = request.form.get('data')
        data = json.loads(data)

        # Generate new time slots based on the dates provided
        new_data = generate_time_slots(fromDate, toDate)
    
        # Update the first three columns of the original data with new data
        for i in range(min(len(data), len(new_data))):
            data[i][:3] = new_data[i]

        data = json.dumps(data)
        
        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)
        role = x['role']

        cursor.execute("select state_name, acronym from states where state_id=%s", (state,))
        state_name = cursor.fetchall()[0][0]

        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})

        file = request.files['excelFile']

        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        if file and allowed_file(file.filename):
            file_name = secure_filename(file.filename)
            from_date = datetime.strptime(fromDate, '%d/%m/%Y').strftime('%d.%m.%Y')
            to_date = datetime.strptime(toDate, '%d/%m/%Y').strftime('%d.%m.%Y')

            directory_path = os.path.join(shared_drive_path,"WEEK_AHEAD_FORECAST_FILES", f"{from_date}-{to_date}", state_name)
            cursor.execute("select * from week_ahead_file_uploads where from_date = to_date(%s, 'DD/MM/YYYY') and to_date = to_date(%s,'DD/MM/YYYY') and state_id = %s", (fromDate, toDate, state))

            existing_revs = cursor.fetchall()

            # Create the directory if it doesn't exist
            os.makedirs(directory_path, exist_ok=True)

            filename = f"{from_date}_{to_date}_{state_name}_rev{len(existing_revs)}.xlsx"
            file_path = os.path.join(directory_path, filename)
            file.save(file_path)

            # Insert data into the database
            cursor.execute("""
                INSERT INTO week_ahead_file_uploads (state_id, from_date, to_date, upload_time, file_name, revision_no, uploaded_by, file_data) 
                VALUES (%s, to_date(%s, 'DD/MM/YYYY'), to_date(%s, 'DD/MM/YYYY'), to_timestamp(%s, 'YYYY-MM-DD HH24:MI:SS'), %s, %s, %s, %s)
            """, (state, fromDate, toDate, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, len(existing_revs), role, data))
            conn.commit()

            # Check if all states have uploaded files for the week
            cursor.execute("""
                SELECT state_id FROM states WHERE state_id NOT IN (
                    SELECT state_id FROM week_ahead_file_uploads WHERE from_date = to_date(%s, 'DD/MM/YYYY') AND to_date = to_date(%s, 'DD/MM/YYYY')
                ) and state_id IN (1,2,3,4,5,7)
            """, (fromDate, toDate))
            pending_states = cursor.fetchall()

            # If all states have uploaded, generate the consolid ated file and send email
            if not pending_states and len(existing_revs) == 0:
                db_data = fetch_max_revision_data(for_week_ahead='week', custom_from_date=fromDate, custom_to_date=toDate)
                if db_data:
                    # file_path = r'D:\Demand Forecasting Page\custom scripts\week_ahead_consolidated.xlsx'
                    file_path_template = os.path.join(shared_drive_path, "FORMATS", "week_ahead_consolidated.xlsx")
                    save_file_path = write_to_excel(file_path_template, db_data, for_week_ahead='week', custom_from_date=datetime.strptime(fromDate, '%d/%m/%Y'), custom_to_date=datetime.strptime(toDate, '%d/%m/%Y'))
                    try:
                        # Send the email with the consolidated file
                        send_mails_list_db = ['mprashad@grid_india.in']
                        cc_mails_list_db = [ 'srldcos@grid_india.in']
                        send_mail_with_attachment(send_mails_list_db, cc_mails_list_db, save_file_path, for_week_ahead='week', custom_from_date=datetime.strptime(fromDate, '%d/%m/%Y'), custom_to_date=datetime.strptime(toDate, '%d/%m/%Y'))
                        print('e')
                        print(f"All states have uploaded. Email with consolidated file sent.")
                    except Exception as email_error:
                        # Handle email failure gracefully
                        print(f"Failed to send email: {email_error}")
                        log_error("Email sending failed", email_error)

            return jsonify({'message': f'Data and file uploaded successfully. Uploaded for Revision-{len(existing_revs)}', "status": "success"})

        else:
            cursor.close()
            return jsonify({'error': 'Invalid file type'})

    except Exception as e:
        log_error("uploadweekahead", e)
        cursor.close()
        return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")


    


@app.route('/api/uploadmonthahead', methods=['POST'])
@jwt_required()
@token_required
def uploadMonthAheadDataAndFile():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # print(request.get_json())
        # print(request["form_data"])
        # print(request["data"])
        header_data = dict(request.headers)
        # print(header_data)
        state = request.form.get('state')
        # print("state",state)
        fromDate = request.form.get('fromDate')

        toDate = request.form.get('toDate')
        # print(disabledDate)
        data = request.form.get('data')
        data = json.loads(data)

        # print(disabledDate, "Date range")

        

        # Generate new time slots based on the dates provided
        # new_data = generate_time_slots(fromDate, toDate)
        # new_data = generate_time_slots(fromDate, toDate)
    
        # # Update the first three columns of the original data with new data
        # for i in range(min(len(data), len(new_data))):
        #     data[i][:3] = new_data[i]


        data = json.dumps(data)
        # print(data[0])

        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)

        username = x['sub']
        # print(username, "username")
        role = x['role']
        

        cursor.execute("select state_name, acronym from states where state_id='{0}'".format(state))

        state_name = cursor.fetchall()[0][0]


            


        # date_string = disabledDate

        # Define the format of the input date string
            # date_format = "%a %b %d %Y %H:%M:%S GMT%z (%Z)"

        # Parse the date string into a datetime object
            # disabledDate = (datetime.strptime(date_string, date_format)).strftime("%Y-%m-%d")

            # print(state, type(disabledDate) )
            # print("Data Received")


        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})

        file = request.files['excelFile']

        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        if file and allowed_file(file.filename):
            file_name = secure_filename(file.filename)

            current_directory = os.getcwd()
            drive_name, path = os.path.splitdrive(current_directory)


            from_date = datetime.strptime(fromDate, '%d/%m/%Y').strftime('%d.%m.%Y')
            to_date = datetime.strptime(toDate, '%d/%m/%Y').strftime('%d.%m.%Y')

            directory_path = os.path.join(shared_drive_path,"MONTH_AHEAD_FORECAST_FILES", from_date+"-"+ to_date, state_name)

            cursor.execute("select * from month_ahead_file_uploads where from_date = to_date('{0}', 'DD/MM/YYYY') and to_date = to_date('{1}','DD/MM/YYYY') and state_id = {2}".format(fromDate,toDate,  state))

            existing_revs = cursor.fetchall()


            # Create the directory if it doesn't exist
            os.makedirs(directory_path, exist_ok=True)

            # Get the current revision number for the state and increment it
            filename = f"{from_date}_{to_date}_{state_name}_rev{len(existing_revs)}.xlsx"

            # Generate the filename based on the current revision number

            file_path = os.path.join(directory_path, filename)

            print(directory_path, "This is the directory path")

            print("file path", file_path)

            # Save the uploaded file in the directory
            if 'excelFile' in request.files:
                file = request.files['excelFile']
                if file.filename != '':
                    # print("entered in save")
                    file.save(file_path)
                    # Insert data into the database
                    cursor.execute("""
                        INSERT INTO month_ahead_file_uploads (state_id, from_date, to_date, upload_time, file_name, revision_no, uploaded_by, file_data) 
                        VALUES (%s, to_date(%s, 'DD/MM/YYYY'), to_date(%s, 'DD/MM/YYYY'), to_timestamp(%s, 'YYYY-MM-DD HH24:MI:SS'), %s, %s, %s, %s)
                    """, (state, fromDate, toDate, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, len(existing_revs), role, data))
                    conn.commit()

                    # Check if all states have uploaded files for the week
                    cursor.execute("""
                        SELECT state_id FROM states WHERE state_id NOT IN (
                            SELECT state_id FROM month_ahead_file_uploads WHERE from_date = to_date(%s, 'DD/MM/YYYY') AND to_date = to_date(%s, 'DD/MM/YYYY')
                        ) and state_id IN (1,2,3,4,5,7)
                    """, (fromDate, toDate))
                    pending_states = cursor.fetchall()

                    # If all states have uploaded, generate the consolid ated file and send email
                    if not pending_states and len(existing_revs) == 0:
                        db_data = fetch_max_revision_data(for_week_ahead='month', custom_from_date=fromDate, custom_to_date=toDate)
                        if db_data:
                            # file_path = r'D:\Demand Forecasting Page\custom scripts\week_ahead_consolidated.xlsx'
                            file_path_template = os.path.join(shared_drive_path, "FORMATS", "month_ahead_consolidated.xlsx")
                            save_file_path = write_to_excel(file_path_template, db_data, for_week_ahead='month', custom_from_date=datetime.strptime(fromDate, '%d/%m/%Y'), custom_to_date=datetime.strptime(toDate, '%d/%m/%Y'))
                            try:
                                # Send the email with the consolidated file
                                send_mails_list_db = ['mprashad@grid_india.in']
                                cc_mails_list_db = [ 'srldcos@grid_india.in']
                                send_mail_with_attachment(send_mails_list_db, cc_mails_list_db, save_file_path, for_week_ahead='month', custom_from_date=datetime.strptime(fromDate, '%d/%m/%Y'), custom_to_date=datetime.strptime(toDate, '%d/%m/%Y'))
                                print('e')
                                print(f"All states have uploaded. Email with consolidated file sent.")
                            except Exception as email_error:
                                # Handle email failure gracefully
                                print(f"Failed to send email: {email_error}")
                                log_error("Email sending failed", email_error)

                    return jsonify({'message': f'Data and file uploaded successfully. Uploaded for Revision-{len(existing_revs)}', "status": "success"})
        else:
            cursor.close()
            return jsonify({'error': 'Invalid file type'})
            
            # return jsonify({'error': 'Invalid file type'})
      
    except Exception as e:
        log_error("uploadmonthahead", e)
        cursor.close()
        return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")
        



@app.route('/api/fetchmonthrevisions', methods=['POST'])
@jwt_required()
@token_required
def fetchMonthRevisions():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        from_date = params["from_date"]
        to_date = params["to_date"]
        state = params["state"]
        cursor.execute("select revision_no from month_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY')".format(params["state"], from_date, to_date))
        revisions_data = cursor.fetchall()
        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for '{0}'-'{1}'".format(from_date, to_date), revisions=revision_list, from_date=from_date, to_date=to_date)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for state {0} for the month '{1}'-'{2}'".format(state_name, from_date, to_date)) 

    except Exception as e:
        log_error("fetchmonthrevisions", e)
        cursor.close()
        return jsonify(message="There is some problem in fetching the revisions! Please contact ERLDC IT", status="failure")
    

@app.route('/api/fetchmonthlyrevisionsdata', methods=['POST'])
@jwt_required()
@token_required
def fetchMonthlyRevisionsData():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)
        cursor.execute("select file_data, from_date, to_date, upload_time, uploaded_by, revision_no from month_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY') and revision_no={3}".format(params["state"], params["from_date"], params["to_date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        from_date = ''
        to_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            from_date = data[0][1].strftime('%d-%b-%Y')
            to_date = data[0][2].strftime('%d-%b-%Y')
            uploaded_time = data[0][3].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][4]
            revision_no = data[0][5]

            cursor.close()
            # print(len(file_data), len(file_data[0]))
            return jsonify(status="success", data=file_data, time=uploaded_time, from_date=from_date,to_date=to_date, revision=revision_no, role=uploaded_by)
        else:
            cursor.close()
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact ERLDC IT!")
            

        # return jsonify(message="Fetched Successfully")

    except Exception as e:
        log_error("fetchmonthlyrevisionsdata", e)
        cursor.close()
        return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")


@app.route('/api/uploadyearahead', methods=['POST'])
@jwt_required()
@token_required
def uploadYearAheadDataAndFile():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # print(request.get_json())
        # print(request["form_data"])
        # print(request["data"])
        header_data = dict(request.headers)
        # print(header_data)
        state = request.form.get('state')
        # print("state",state)
        fromDate = request.form.get('fromDate')

        toDate = request.form.get('toDate')
        # print(disabledDate)
        data = request.form.get('data')
        data = json.loads(data)

        # print(disabledDate, "Date range")

        data = json.dumps(data)
        # print(data[0])

        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)

        username = x['sub']
        # print(username, "username")
        role = x['role']
        

        cursor.execute("select state_name, acronym from states where state_id='{0}'".format(state))

        state_name = cursor.fetchall()[0][0]


            


        # date_string = disabledDate

    # Define the format of the input date string
        # date_format = "%a %b %d %Y %H:%M:%S GMT%z (%Z)"

    # Parse the date string into a datetime object
        # disabledDate = (datetime.strptime(date_string, date_format)).strftime("%Y-%m-%d")

        # print(state, type(disabledDate) )
        # print("Data Received")


        if 'excelFile' not in request.files:
            return jsonify({'error': 'No file part'})

        file = request.files['excelFile']

        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        if file and allowed_file(file.filename):
            file_name = secure_filename(file.filename)

            current_directory = os.getcwd()
            drive_name, path = os.path.splitdrive(current_directory)

            print(drive_name)

            from_date = datetime.strptime(fromDate, '%d/%m/%Y').strftime('%d.%m.%Y')
            to_date = datetime.strptime(toDate, '%d/%m/%Y').strftime('%d.%m.%Y')

            directory_path = os.path.join(shared_drive_path,"YEAR_AHEAD_FORECAST_FILES", from_date+"-"+ to_date, state_name)

            cursor.execute("select * from year_ahead_file_uploads where from_date = to_date('{0}', 'DD/MM/YYYY') and to_date = to_date('{1}','DD/MM/YYYY') and state_id = {2}".format(fromDate,toDate,  state))

            existing_revs = cursor.fetchall()


            # Create the directory if it doesn't exist
            os.makedirs(directory_path, exist_ok=True)

            # Get the current revision number for the state and increment it
            filename = f"{from_date}_{to_date}_{state_name}_rev{len(existing_revs)}.xlsx"

            # Generate the filename based on the current revision number

            file_path = os.path.join(directory_path, filename)

            # print(directory_path, "This is the directory path")

            # print("file path", file_path)

            # Save the uploaded file in the directory
            if 'excelFile' in request.files:
                file = request.files['excelFile']
                if file.filename != '':
                    # print("entered in save")
                    file.save(file_path)
                    if len(existing_revs) > 0:
                        cursor.execute("insert into year_ahead_file_uploads (state_id, from_date,to_date, upload_time, file_name, revision_no, uploaded_by, file_data) values({0}, to_date('{1}', 'DD/MM/YYYY'),to_date('{2}','DD/MM/YYYY'), to_timestamp('{3}', 'YYYY-MM-DD HH24:MI:SS'), '{4}', {5}, '{6}', '{7}')".format(state, fromDate, toDate, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, len(existing_revs), role, data))
                        conn.commit()
                        cursor.close()
                        # cursor.execute("update file_contents set file_data =  '{0}' where state_id = {1} and upload_date = to_date('{2}', 'YYYY-MM-DD')".format( data, state, disabledDate))
                        return jsonify({'message': 'Data and file uploaded successfully. Uploaded with Revision-{0}'.format(len(existing_revs)), "status": "success"})

                    else:
                        cursor.execute("insert into year_ahead_file_uploads (state_id, from_date,to_date, upload_time, file_name, revision_no, uploaded_by, file_data) values({0}, to_date('{1}', 'DD/MM/YYYY'),to_date('{2}','DD/MM/YYYY'), to_timestamp('{3}', 'YYYY-MM-DD HH24:MI:SS'), '{4}', {5}, '{6}', '{7}')".format(state, fromDate,toDate, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, len(existing_revs), role, data))
                        conn.commit()
                        cursor.close()
                        # cursor.execute("insert into file_contents (state_id, upload_date, file_data) values({0}, to_date('{1}', 'YYYY-MM-DD'), '{2}')".format(state, disabledDate, data))
                        return jsonify({'message': 'Data and file uploaded successfully. Uploaded with Revision-{0}'.format(len(existing_revs)), "status":"success"})
            # file.save("D:\\forecast_excel_store\\"+file_name)
            # print("file saved successfully")
            # Process the form data and uploaded file as needed
            # You can access 'name' and 'email' here

            # return jsonify({'message': 'Data and file uploaded successfully'})
        else:
            return jsonify({'error': 'Invalid file type'})

    except Exception as e:
        log_error("uploadyearahead", e)
        cursor.close()
        return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")





@app.route('/api/fetchyearlyrevisionsdata', methods=['POST'])
@jwt_required()
@token_required
def fetchYearlyRevisionsData():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)
        cursor.execute("select file_data, from_date, to_date, upload_time, uploaded_by, revision_no from year_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY') and revision_no={3}".format(params["state"], params["from_date"], params["to_date"], int(params["revision"])))
        data = cursor.fetchall()

        file_data = []
        uploaded_time = ''
        from_date = ''
        to_date = ''
        uploaded_by = ''
        # role = ''
        revision_no = int()



        if len(data) > 0:
            file_data = data[0][0]
            from_date = data[0][1].strftime('%d-%b-%Y')
            to_date = data[0][2].strftime('%d-%b-%Y')
            uploaded_time = data[0][3].strftime('%d-%b-%Y %H:%M:%S %p')
            uploaded_by = data[0][4]
            revision_no = data[0][5]
            cursor.close()
            # print(len(file_data), len(file_data[0]))
            # cursor.c
            return jsonify(status="success", data=file_data, time=uploaded_time, from_date=from_date,to_date=to_date, revision=revision_no, role=uploaded_by)
            
        else:
            return jsonify(status="failure", message="There is a Problem in fetching the data, Please contact ERLDC IT!")
            

        return jsonify(message="Fetched Successfully")
    
    except Exception as e:
        log_error("fetchyearlyrevisionsdata", e)
        cursor.close()
        return jsonify(message="There is some problem in fetching the revisions data! Please contact ERLDC IT", status="failure")



@app.route('/api/fetchyearrevisions', methods=['POST'])
@jwt_required()
@token_required
def fetchYearRevisions():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        from_date = params["from_date"]
        to_date = params["to_date"]
        state = params["state"]
        cursor.execute("select revision_no from year_ahead_file_uploads where state_id = {0} and from_date = to_date('{1}', 'DD/MM/YYYY') and to_date=to_date('{2}', 'DD/MM/YYYY')".format(params["state"], from_date, to_date))
        revisions_data = cursor.fetchall()
        revision_list = [i[0] for i in revisions_data]

        cursor.execute("select state_name from states where state_id = {0}".format(state))
        state_name = cursor.fetchall()[0][0]


        if len(revision_list) > 0:
            cursor.close()
            return jsonify(status="success", message="Fetched Successfully for '{0}'-'{1}'".format(from_date, to_date), revisions=revision_list, from_date=from_date, to_date=to_date)
        else:
            cursor.close()
            return jsonify(status="failure", message="There are no Uploads for state {0} for the year '{1}'-'{2}'".format(state_name, from_date, to_date))    
    
    except Exception as e:
        log_error("fetchyearrevisions", e)
        cursor.close()
        return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")






@app.route('/api/uploadstatus')
@jwt_required()
@token_required
@session_token_required
def scatterPlotUploadStatus():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Day Ahead Data Processing
        header_data = dict(request.headers)
        state = request.form.get('state')

        token = header_data['Authorization'].split()[1]
        x = decode_token(token, csrf_value=None, allow_expired=False)


        state_id_query = """select state_id from public.states where username = %s"""
        cursor.execute(state_id_query, (x['sub'],))
        state_id = cursor.fetchall()[0][0]

        # print(state)
        end_date = datetime.now() + timedelta(days=1)
        start_date = end_date - timedelta(days=30)

        sql_query = """
                    WITH min_revision AS (
                    SELECT 
                        state_id, 
                        upload_date, 
                        MIN(revision_no) AS min_revision_no
                    FROM 
                        file_uploads
                    GROUP BY 
                        state_id, 
                        upload_date
                )
                SELECT 
                    states.state_name,
                    COALESCE(file_uploads.upload_date, %s) AS upload_date,
                    COALESCE(min_uploads.upload_time, NULL) AS upload_time,
                    CASE
                        WHEN min_uploads.upload_time IS NULL THEN 2  -- Not Uploaded
                        WHEN min_uploads.upload_time < (file_uploads.upload_date - INTERVAL '1 day' + INTERVAL '10 hours') THEN 1  -- Proper Upload before 10 AM on the previous day
                        ELSE 0  -- Late Upload
                    END AS upload_status_code,
                    COUNT(file_uploads.state_id) AS upload_count
                FROM 
                    states
                LEFT JOIN 
                    file_uploads ON states.state_id = file_uploads.state_id
                    AND file_uploads.upload_date BETWEEN %s AND %s
                LEFT JOIN (
                    SELECT 
                        file_uploads.state_id, 
                        file_uploads.upload_date, 
                        file_uploads.upload_time
                    FROM 
                        file_uploads
                    INNER JOIN min_revision ON 
                        file_uploads.state_id = min_revision.state_id
                        AND file_uploads.upload_date = min_revision.upload_date
                        AND file_uploads.revision_no = min_revision.min_revision_no
                ) AS min_uploads ON 
                    file_uploads.state_id = min_uploads.state_id 
                    AND file_uploads.upload_date = min_uploads.upload_date
                WHERE 
                    states.state_id IN (1,2,3,4,5,7)
                GROUP BY 
                    states.state_name, 
                    file_uploads.upload_date,
                    min_uploads.upload_time  
                ORDER BY
                    states.state_name,
                    upload_date DESC;
        """
        cursor.execute(sql_query, (end_date, start_date, end_date))
        results = cursor.fetchall()

        day_data = []
        date_range = [end_date - timedelta(days=i) for i in range(30)]
        state_names = set(result[0] for result in results)

        for state_name in state_names:
            state_data = {"name": state_name, "data": []}
            for date in date_range:
                date_str = date.strftime('%Y-%m-%d')
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1].strftime("%Y-%m-%d") == date_str:
                        upload_count = result[4]
                        upload_time = result[2].strftime("%Y-%m-%d %H:%M:%S") if result[2] is not None else None
                        upload_status_code = result[3]
                        found_result = True
                        break
                if not found_result:
                    upload_count = 0
                    upload_time = None
                    upload_status_code = 2  # Not Uploaded
                state_data["data"].append({
                    'x': date_str, 
                    'y': upload_status_code, 
                    'upload_time': upload_time, 
                    'upload_count': upload_count
                })
            day_data.append(state_data)

        day_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }

        mape_comp_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%d/%m/%Y'),
            "end_date": end_date.strftime('%d/%m/%Y')
        }


        # for i in day_data:
        #     print(i["name"])

        #######################################################################################################################################
        #######################################################################################################################################
        # Week Ahead Data Processing
                


        # Week Ahead Data Processing
        # Calculate date ranges
        today = datetime.now()
        # Get the start of this week (Monday)
        this_week_start = today - timedelta(days=today.weekday())

        # Setting date ranges
        start_date = this_week_start - timedelta(weeks=3)  # Start date two weeks before this week
        end_date = this_week_start + timedelta(weeks=2) - timedelta(days=1)  # End date two weeks after this week

        # Adjusting end_date to ensure it ends on the last day of the intended week
        if end_date.weekday() != 6:  # Check if end_date is not a Sunday
            end_date = end_date - timedelta(days=end_date.weekday() + 1)  # Adjust to the last Sunday before or on end_date

        sql_query = """
            WITH min_revision AS (
                SELECT 
                    state_id, 
                    from_date, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    week_ahead_file_uploads
                WHERE
                    from_date BETWEEN %s AND %s
                GROUP BY 
                    state_id, 
                    from_date
            ), min_uploads AS (
                SELECT 
                    fu.state_id, 
                    fu.from_date, 
                    fu.upload_time,
                    fu.revision_no
                FROM 
                    week_ahead_file_uploads fu
                INNER JOIN min_revision mr ON
                    fu.state_id = mr.state_id
                    AND fu.from_date = mr.from_date
                    AND fu.revision_no = mr.min_revision_no
            )
            SELECT 
                states.state_name,
                COALESCE(mu.from_date, %s) AS week_start_date,
                COALESCE(mu.upload_time, NULL) AS upload_time,
                CASE
                    WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                    WHEN mu.upload_time < DATE_TRUNC('week', mu.from_date - INTERVAL '1 week')  + INTERVAL '1 day' THEN 1  -- Uploaded on time, before the first working day of the week
                    ELSE 0  -- Late Upload
                END AS upload_status_code,
                COUNT(mu.state_id) AS upload_count
            FROM 
                states
            LEFT JOIN 
                min_uploads mu ON states.state_id = mu.state_id
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                mu.from_date,
                mu.upload_time  
            ORDER BY
                states.state_name,
                mu.from_date DESC;
        """
        cursor.execute(sql_query, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        results = cursor.fetchall()


        cursor.execute("SELECT state_name, state_id FROM states WHERE state_id IN (1,2,3,4,5,7)")
        states_list = cursor.fetchall()
        all_states = [row[0] for row in states_list]


        state_dict = {row[0]: row[1] for row in states_list}


        week_range = sorted([(start_date + timedelta(weeks=i)).date() for i in range(5)], reverse=True)  # Generate Mondays for 6 weeks

        week_data = []
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for week_start in week_range:
                week_start_date = week_start
                week_end_date = week_start_date + timedelta(days=6)  # End on Sunday
                week_range_str = f"{week_start_date.strftime('%Y-%m-%d')} to {week_end_date.strftime('%Y-%m-%d')}"

                upload_count = 0
                upload_status_code = 2  # Default to Not Uploaded
                upload_time = None
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1] == week_start_date:
                        upload_count = result[4]
                        upload_time = result[2].strftime("%Y-%m-%d %H:%M:%S") if result[2] is not None else None
                        upload_status_code = result[3]
                        found_result = True
                        break

                if not found_result:
                    upload_count = 0
                    upload_status_code = 2  # Not Uploaded
                state_data["data"].append({'x': week_range_str, 'y': upload_status_code, 'upload_time': upload_time, 'upload_count': upload_count})
            week_data.append(state_data)

        week_dates = {
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }






        #######################################################################################################################################
        #######################################################################################################################################
        # Month Ahead Data Processing
        # Initialize current date and calculate relevant month starts
        today = datetime.now()
        current_month_start = today.replace(day=1)
        previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        next_month_start = (current_month_start + timedelta(days=31)).replace(day=1)

        # Define the date ranges
        start_date = previous_month_start
        end_date = next_month_start.replace(day=monthrange(next_month_start.year, next_month_start.month)[1])

        # Prepare the SQL query to fetch the data
        sql_query = """
        WITH min_revision AS (
                SELECT 
                    state_id, 
                    DATE_TRUNC('month', from_date) AS from_month, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    month_ahead_file_uploads
                WHERE
                    from_date BETWEEN %s AND %s
                GROUP BY 
                    state_id, 
                    from_month
            ), min_uploads AS (
                SELECT 
                    fu.state_id, 
                    DATE_TRUNC('month', fu.from_date) AS from_month, 
                    fu.upload_time,
                    fu.revision_no
                FROM 
                    month_ahead_file_uploads fu
                INNER JOIN min_revision mr ON
                    fu.state_id = mr.state_id
                    AND DATE_TRUNC('month', fu.from_date) = mr.from_month
                    AND fu.revision_no = mr.min_revision_no
            )
            SELECT 
                states.state_name,
                COALESCE(mu.from_month, %s) AS month_start_date,
                COALESCE(mu.upload_time, NULL) AS upload_time,
                CASE
                    WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                    WHEN mu.upload_time < DATE_TRUNC('month', mu.from_month) - INTERVAL '1 month' + INTERVAL '5 day' THEN 1  -- Uploaded on time
                    ELSE 0  -- Late Upload
                END AS upload_status_code,
                COUNT(mu.state_id) AS upload_count
            FROM 
                states
            LEFT JOIN 
                min_uploads mu ON states.state_id = mu.state_id
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                mu.from_month,
                mu.upload_time  
            ORDER BY
                states.state_name,
                mu.from_month DESC;
        """
        cursor.execute(sql_query, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        results = cursor.fetchall()

        # Fetch the state names from the database
        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        # Initialize month range and data structure
        month_range = [next_month_start, current_month_start, previous_month_start]
        month_data = []

        # Process results and sort by latest month first
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for month_start in sorted(month_range, reverse=True):
                _, last_day_of_month = monthrange(month_start.year, month_start.month)
                month_end = month_start.replace(day=last_day_of_month)
                month_range_str = f"{month_start.strftime('%Y-%m-%d')} to {month_end.strftime('%Y-%m-%d')}"

                upload_count = 0
                upload_status_code = 2  # Default to Not Uploaded
                upload_time = None
                found_result = False

                # Check each result for the current state and month
                for result in results:
                    result_state, result_month_start, result_upload_time, result_status, result_count = result
                    if result_state == state_name and result_month_start.strftime('%Y-%m-%d') == month_start.strftime('%Y-%m-%d'):
                        upload_count = result_count
                        upload_time = result_upload_time.strftime("%Y-%m-%d %H:%M:%S") if result_upload_time else None
                        upload_status_code = result_status
                        found_result = True
                        break

                if not found_result:
                    upload_count = 0
                    upload_status_code = 2  # Not Uploaded

                state_data["data"].append({
                    'x': month_range_str,
                    'y': upload_status_code,
                    'upload_time': upload_time,
                    'upload_count': upload_count
                })

            month_data.append(state_data)

        # Prepare JSON for frontend
        month_dates = {
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }


        ############### Year Ahead Forecast Status

        # Financial year calculation
        
        today = datetime.now()

        # Determine the current and next financial years
        if today.month < 4:
            current_financial_year_start = datetime(today.year - 1, 4, 1)
        else:
            current_financial_year_start = datetime(today.year, 4, 1)
        current_financial_year_end = datetime(current_financial_year_start.year + 1, 3, 31)
        next_financial_year_start = datetime(current_financial_year_start.year + 1, 4, 1)
        next_financial_year_end = datetime(next_financial_year_start.year + 1, 3, 31)

        # SQL query to fetch relevant data for both financial years
        sql_query = """
        SELECT 
            states.state_name,
            yafu.from_date,
            yafu.to_date,
            yafu.upload_time,
            CASE
                WHEN yafu.upload_time IS NULL THEN 2
                WHEN yafu.upload_time <= yafu.from_date + INTERVAL '5 months' - INTERVAL '1 day' THEN 1
                ELSE 0
            END AS upload_status,
            COUNT(yafu.state_id) AS upload_count
        FROM 
            states
        LEFT JOIN 
            year_ahead_file_uploads yafu ON states.state_id = yafu.state_id
        WHERE 
            states.state_id IN (1,2,3,4,5,7)
            AND (yafu.from_date BETWEEN %s AND %s OR yafu.from_date IS NULL)
        GROUP BY 
            states.state_name, yafu.from_date, yafu.to_date, yafu.upload_time
        ORDER BY
            states.state_name, yafu.from_date DESC;
        """
        cursor.execute(sql_query, (current_financial_year_start.strftime('%Y-%m-%d'), next_financial_year_end.strftime('%Y-%m-%d')))
        results = cursor.fetchall()

        # Fetch state names to ensure coverage of all states
        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        # Organize the fetched data by state and financial year
        year_data = []
        financial_years = [next_financial_year_start,current_financial_year_start]

        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for financial_year in financial_years:
                found = False
                for result in results:
                    result_state, from_date, to_date, upload_time, upload_status, upload_count = result
                    if result_state == state_name and from_date and from_date == financial_year.date():
                        found = True
                        upload_time_formatted = upload_time.strftime("%Y-%m-%d %H:%M:%S") if upload_time else None
                        state_data["data"].append({
                            'x': f"{from_date.strftime('%Y-%m-%d')} to {to_date.strftime('%Y-%m-%d')}",
                            'y': upload_status,
                            'upload_time': upload_time_formatted,
                            'upload_count': upload_count
                        })
                        break
                if not found:
                    state_data["data"].append({
                        'x': f"{financial_year.strftime('%Y-%m-%d')} to {financial_year.replace(year=financial_year.year + 1, month=3, day=31).strftime('%Y-%m-%d')}",
                        'y': 2,
                        'upload_time': None,
                        'upload_count': 0
                    })
            year_data.append(state_data)

        # Prepare JSON data for the frontend
        year_dates = {
            "start_date": current_financial_year_start.strftime('%Y-%m-%d'),
            "end_date": next_financial_year_end.strftime('%Y-%m-%d')
        }




        # url = host_url +"/mapechart"
        # headers = {
        #     'Authorization': request.headers.get('Authorization'),
        #     'Content-Type': 'application/json'
        # }
        # payload = {
        #     "params": {
        #         "fromDate": from_date,
        #         "toDate": to_date,
        #         "state": state_id
        #     }
        # }

        # print(header_data)

        # token = header_data['Authorization'].split()[1]
        # x = decode_token(token, csrf_value=None, allow_expired=False)

        # username = x['sub']
        # # print(username, "username")
        # # role = x['role']

        # # print(state_dict)

        # if username in state_dict.keys():
        #     url = host_url +"/mapechart"
        #     headers = {
        #         'Authorization': request.headers.get('Authorization'),
        #         'Content-Type': 'application/json'
        #     }
        #     payload = {
        #         "params": {
        #             "fromDate": mape_comp_dates['start_date'],
        #             "toDate": mape_comp_dates['end_date'],
        #             "state": state_dict[username]
        #         }
        #     }



        #     response = requests.post(url, json=payload, headers=headers)
        #     cursor.close()

        #     if response.status_code == 200:
        #         # pass
        #         # print(response.json().keys())
        #         res_data = response.json()
        #         print(res_data.keys())
        #         print(res_data['title'])
        #         if res_data['status'] == 'success':    
        #             return jsonify(day=day_data, week=week_data, month=month_data, year = year_data, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates, mape_data= res_data['data'],mape_title = res_data['title'],comp_data = res_data['comp_data'],  status="success")
        #         else:
        #             return jsonify(day=day_data, week=week_data, month=month_data, year = year_data, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates,  status="success")

        #     else:
        #         print("Response not recieved")
        #         return jsonify(day=day_data, week=week_data, month=month_data, year = year_data, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates,  status="success")







        


        return jsonify(day=day_data, week=week_data, month=month_data, year = year_data, day_dates=day_dates, week_dates=week_dates, month_dates=month_dates, year_dates=year_dates, status="success")
    
    except Exception as e:
        # print(error)
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print(exc_type, fname, exc_tb.tb_lineno)
        log_error("uploadstatus", e)
        # cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact ERLDC IT!", status="failure")











@app.route('/api/mapechart', methods=['POST'])
@jwt_required()
@token_required
@session_token_required
def mapeChart():
    try:
        conn = psycopg2.connect(
        database="demand_forecast_states", user='ra_admin', 
        password='admin', host='localhost', port='5432'
        )
        cursor = conn.cursor()
        params = request.get_json()
        # print(params)

        from_date = params["params"]["fromDate"]
        to_date = params["params"]["toDate"]  
        state_id = int(params["params"]["state"])

        # print(state_id, type(state_id))

        cursor.execute("select state_name from states where state_id = {0}".format(state_id))
        state_name = cursor.fetchall()[0][0]

        region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
        selected_state_id = state_id  # Set to the desired single state or region (use 6 for the entire region)

        # Define expected state IDs based on selected_state_id
        expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

        date_range = pd.date_range(start=datetime.strptime(from_date, "%d/%m/%Y"), end=datetime.strptime(to_date, "%d/%m/%Y"))
        date_df = pd.DataFrame(date_range, columns=["D_F_DATE"])



        # Fetch actual data and filter by expected state IDs
        cursor.execute('''
            select s.state_id, process_date, demand_met, scada_demand 
            from states s 
            join actual_demand ad on s.state_id = ad.state_id 
            where process_date between to_date('{0}', 'DD/MM/YYYY') and to_date('{1}', 'DD/MM/YYYY')
            order by process_date
        '''.format(from_date, to_date))

        sr_actual_data_result = cursor.fetchall()

        # Aggregate actual data by date for selected state or region
        daily_totals = defaultdict(lambda: {"DAY_ACTUAL": [0] * 96, "SCADA_DEMAND": [0] * 96, "STATE_IDS": set()})
        for record in sr_actual_data_result:
            state_id, process_date, demand_met_list, scada_demand_list = record
            if state_id not in expected_state_ids:  # Filter out any unexpected states
                continue
            date_str = process_date.strftime("%Y-%m-%d")

            if demand_met_list is None:
                demand_met_list = [0] * 96
            if scada_demand_list is None:
                scada_demand_list = [0] * 96
            
            daily_totals[date_str]["STATE_IDS"].add(state_id)
            
            if state_id == 3 and (not demand_met_list or all(x == 0 for x in demand_met_list)):
                demand_met_list = scada_demand_list

            daily_totals[date_str]["DAY_ACTUAL"] = [x + y for x, y in zip(daily_totals[date_str]["DAY_ACTUAL"], demand_met_list)]
            daily_totals[date_str]["SCADA_DEMAND"] = [x + y for x, y in zip(daily_totals[date_str]["SCADA_DEMAND"], scada_demand_list)]

        sr_day_actual_json_list = [
            {"A_DATE": date, "DAY_ACTUAL": totals["DAY_ACTUAL"], "SCADA_DEMAND": totals["SCADA_DEMAND"], "STATE_IDS": list(totals["STATE_IDS"])}
            for date, totals in daily_totals.items()
        ]
        sr_actual_day_df = pd.DataFrame(sr_day_actual_json_list)


        # print(sr_actual_day_df['A_DATE'].unique())

        def intraday_mape(from_date, to_date, sr_actual_day_df):
            cursor.execute('''
                WITH RECURSIVE DateRange AS (
                    SELECT to_date('{0}', 'DD/MM/YYYY')::date AS upload_date
                    UNION ALL
                    SELECT (upload_date + INTERVAL '1 day')::date
                    FROM DateRange
                    WHERE upload_date + INTERVAL '1 day' <= to_date('{1}', 'DD/MM/YYYY')
                ),
                MaxRevisions AS (
                    SELECT state_id, upload_date, MAX(revision_no) AS max_revision
                    FROM intraday_file_uploads
                    WHERE upload_date BETWEEN to_date('{0}', 'DD/MM/YYYY') AND to_date('{1}', 'DD/MM/YYYY')
                    GROUP BY state_id, upload_date
                )
                SELECT dr.upload_date, t.state_id, t.revision_no, t.file_data
                FROM DateRange dr
                LEFT JOIN intraday_file_uploads t
                    ON dr.upload_date = t.upload_date
                    AND t.revision_no = (SELECT max_revision FROM MaxRevisions mr WHERE mr.state_id = t.state_id AND mr.upload_date = t.upload_date)
                ORDER BY dr.upload_date, t.state_id;
            '''.format(from_date, to_date))

            sr_day_data = cursor.fetchall()

            # Aggregate forecast data by date for selected state or region
            sr_day_forecast_dict = defaultdict(lambda: {"INTRADAY_FORECAST": [0] * 96, "STATE_IDS": set()})
            for record in sr_day_data:
                upload_date, state_id, revision_no, file_data = record
                if state_id not in expected_state_ids:
                    continue
                date_str = upload_date.strftime("%Y-%m-%d")
                
                if file_data:
                    day_forecast_values = [forecast[2] for forecast in file_data]
                    sr_day_forecast_dict[date_str]["STATE_IDS"].add(state_id)
                    sr_day_forecast_dict[date_str]["INTRADAY_FORECAST"] = [x + y for x, y in zip(sr_day_forecast_dict[date_str]["INTRADAY_FORECAST"], day_forecast_values)]

            sr_day_forecast_json_list = [
                {"D_F_DATE": date, "INTRADAY_FORECAST": data["INTRADAY_FORECAST"], "STATE_IDS": list(data["STATE_IDS"])}
                for date, data in sr_day_forecast_dict.items()
            ]
            # Create the DataFrame
            if sr_day_forecast_json_list:  # If the list is not empty
                sr_forecast_day_df = pd.DataFrame(sr_day_forecast_json_list)
            else:  # If the list is empty, create an empty DataFrame with required columns
                sr_forecast_day_df = pd.DataFrame(columns=["D_F_DATE", "INTRADAY_FORECAST", "STATE_IDS"])

            formatted_forecast_data, formatted_actual_data = [], []
            sr_mape_dict = {"name": "Intraday", "data": []}

            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            if "D_F_DATE" in sr_forecast_day_df.columns:
                sr_forecast_day_df["D_F_DATE"] = pd.to_datetime(sr_forecast_day_df["D_F_DATE"])
            sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_forecast_day_df, on="D_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            for _, row in sr_actual_day_df.iterrows():
                day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                base_date = row['A_DATE']
                for i, actual_value in enumerate(day_actual_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(actual_value)
                        })




            for _, row in sr_result_df.iterrows():
                temp = {"x": row["A_DATE"].strftime("%Y-%m-%d")}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()
                forecast_day_values = row["INTRADAY_FORECAST"]
                if forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                elif (len(row["INTRADAY_FORECAST"]) != 96 or pd.isna(row["INTRADAY_FORECAST"]).any()):
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                else:
                    mape_value = calculate_mape(day_actual_values, row["INTRADAY_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None
                
                    # Populate formatted_actual_data and formatted_forecast_data
                    base_date = row['D_F_DATE']
                    # print(row["A_DATE"])
                    forecast_day_values = row["INTRADAY_FORECAST"]
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })
                    
                    
                sr_mape_dict["data"].append(temp)

            return formatted_forecast_data, sr_mape_dict
        
        def day_ahead_mape(from_date, to_date, sr_actual_day_df):
            cursor.execute('''
                WITH RECURSIVE DateRange AS (
                    SELECT to_date('{0}', 'DD/MM/YYYY')::date AS upload_date
                    UNION ALL
                    SELECT (upload_date + INTERVAL '1 day')::date
                    FROM DateRange
                    WHERE upload_date + INTERVAL '1 day' <= to_date('{1}', 'DD/MM/YYYY')
                ),
                MaxRevisions AS (
                    SELECT state_id, upload_date, MAX(revision_no) AS max_revision
                    FROM file_uploads
                    WHERE upload_date BETWEEN to_date('{0}', 'DD/MM/YYYY') AND to_date('{1}', 'DD/MM/YYYY')
                    GROUP BY state_id, upload_date
                )
                SELECT dr.upload_date, t.state_id, t.revision_no, t.file_data
                FROM DateRange dr
                LEFT JOIN file_uploads t
                    ON dr.upload_date = t.upload_date
                    AND t.revision_no = (SELECT max_revision FROM MaxRevisions mr WHERE mr.state_id = t.state_id AND mr.upload_date = t.upload_date)
                ORDER BY dr.upload_date, t.state_id;
            '''.format(from_date, to_date))

            sr_day_data = cursor.fetchall()

            # Aggregate forecast data by date for selected state or region
            sr_day_forecast_dict = defaultdict(lambda: {"DAY_FORECAST": [0] * 96, "STATE_IDS": set()})
            for record in sr_day_data:
                upload_date, state_id, revision_no, file_data = record
                if state_id not in expected_state_ids:
                    continue
                date_str = upload_date.strftime("%Y-%m-%d")
                
                if file_data:
                    day_forecast_values = [forecast[2] for forecast in file_data]
                    sr_day_forecast_dict[date_str]["STATE_IDS"].add(state_id)
                    sr_day_forecast_dict[date_str]["DAY_FORECAST"] = [x + y for x, y in zip(sr_day_forecast_dict[date_str]["DAY_FORECAST"], day_forecast_values)]

            sr_day_forecast_json_list = [
                {"D_F_DATE": date, "DAY_FORECAST": data["DAY_FORECAST"], "STATE_IDS": list(data["STATE_IDS"])}
                for date, data in sr_day_forecast_dict.items()
            ]
            sr_forecast_day_df = pd.DataFrame(sr_day_forecast_json_list)

            formatted_forecast_data, formatted_actual_data = [], []
            sr_mape_dict = {"name": "Day Ahead", "data": []}

            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            sr_forecast_day_df["D_F_DATE"] = pd.to_datetime(sr_forecast_day_df["D_F_DATE"])
            sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_forecast_day_df, on="D_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            for _, row in sr_actual_day_df.iterrows():
                day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                base_date = row['A_DATE']
                for i, actual_value in enumerate(day_actual_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_actual_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(actual_value)
                        })




            for _, row in sr_result_df.iterrows():
                temp = {"x": row["A_DATE"].strftime("%Y-%m-%d")}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()
                forecast_day_values = row["DAY_FORECAST"]
                if forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                elif (len(row["DAY_FORECAST"]) != 96 or pd.isna(row["DAY_FORECAST"]).any()):
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                    temp["y"] = None
                else:
                    
                    mape_value = calculate_mape(day_actual_values, row["DAY_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None
                
                    # Populate formatted_actual_data and formatted_forecast_data
                    base_date = row['D_F_DATE']
                    # print(row["A_DATE"])
                    forecast_day_values = row["DAY_FORECAST"]
                    for i, forecast_value in enumerate(forecast_day_values):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_forecast_data.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })
                    
                    
                sr_mape_dict["data"].append(temp)

            return formatted_forecast_data, formatted_actual_data, sr_mape_dict

        def week_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id):
            # Define expected state IDs based on selected_state_id
            region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
            expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

            cursor.execute('''
                WITH MaxRevisions AS (
                    SELECT
                        state_id,
                        from_date,
                        to_date,
                        MAX(revision_no) AS max_revision
                    FROM week_ahead_file_uploads
                    WHERE
                        (from_date <= to_date('{1}', 'DD/MM/YYYY') AND to_date >= to_date('{0}', 'DD/MM/YYYY'))
                    GROUP BY state_id, from_date, to_date
                )
                SELECT
                    t.state_id,
                    t.from_date,
                    t.to_date,
                    t.revision_no,
                    t.file_data
                FROM week_ahead_file_uploads t
                INNER JOIN MaxRevisions mr
                    ON t.state_id = mr.state_id
                    AND t.from_date = mr.from_date
                    AND t.to_date = mr.to_date
                    AND t.revision_no = mr.max_revision
                WHERE
                    t.state_id IN ({2})
                ORDER BY t.from_date;
            '''.format(from_date, to_date, ','.join(map(str, expected_state_ids))))

            sr_week_data = cursor.fetchall()

            # Expand week-ahead forecast data by date range for each state
            expanded_week_forecast = []
            for record in sr_week_data:
                state_id, from_date, to_date, _, file_data = record
                if state_id not in expected_state_ids:
                    continue
                if file_data:
                    week_forecast_values = [forecast[3] for forecast in file_data]

                    current_date = from_date
                    while current_date <= to_date:
                        forecast_index = (current_date - from_date).days * 96  # Start index for 96 intervals
                        daily_forecast_96 = week_forecast_values[forecast_index:forecast_index + 96]
                        
                        if len(daily_forecast_96) == 96:
                            expanded_week_forecast.append({
                                "W_F_DATE": current_date.strftime("%Y-%m-%d"),
                                "WEEK_FORECAST": daily_forecast_96,
                                "STATE_ID": state_id
                            })
                        else:
                            print(f"Warning: Insufficient forecast data for {current_date.strftime('%Y-%m-%d')}")
                        
                        current_date += timedelta(days=1)

            # Convert expanded forecast data to DataFrame
            sr_expanded_week_forecast_df = pd.DataFrame(expanded_week_forecast)

            # Group by date and sum forecasts across states for each day
            sr_week_forecast_daily = (
                sr_expanded_week_forecast_df.groupby("W_F_DATE").apply(
                    lambda group: {
                        "WEEK_FORECAST": np.sum(np.vstack(group["WEEK_FORECAST"]), axis=0).tolist(),
                        "STATE_IDS": list(group["STATE_ID"])
                    }
                )
            ).apply(pd.Series).reset_index()

            formatted_week_forecast = []

            # Merge data frames
            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            sr_week_forecast_daily["W_F_DATE"] = pd.to_datetime(sr_week_forecast_daily["W_F_DATE"])
            sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_week_forecast_daily, left_on="D_F_DATE", right_on="W_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # Calculate MAPE with complete state data only
            sr_mape_dict = {"name": "Week Ahead", "data": []}
            for _, row in sr_result_df.iterrows():
                temp = {"x": row["A_DATE"].strftime('%Y-%m-%d')}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()

                if forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_week_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                elif (len(row["WEEK_FORECAST"]) != 96 or pd.isna(row["WEEK_FORECAST"]).any()):
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_week_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                else:
                    day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                    mape_value = calculate_mape(day_actual_values, row["WEEK_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None

                    base_date = row['A_DATE']
                    for i, forecast_value in enumerate(row["WEEK_FORECAST"]):
                        timestamp = row['D_F_DATE'] + timedelta(minutes=15 * i)
                        formatted_week_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })

                sr_mape_dict["data"].append(temp)

            return formatted_week_forecast, sr_mape_dict

        def month_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id):
            # Define expected state IDs based on selected_state_id
            region_state_ids = {1, 2, 3, 4, 5, 7}  # IDs for the Southern Region
            expected_state_ids = region_state_ids if selected_state_id == 6 else {selected_state_id}

            cursor.execute('''
                WITH MaxRevisions AS (
                    SELECT
                        state_id,
                        from_date,
                        to_date,
                        MAX(revision_no) AS max_revision
                    FROM month_ahead_file_uploads
                    WHERE
                        (from_date <= to_date('{1}', 'DD/MM/YYYY') AND to_date >= to_date('{0}', 'DD/MM/YYYY'))
                    GROUP BY state_id, from_date, to_date
                )
                SELECT
                    t.state_id,
                    t.from_date,
                    t.to_date,
                    t.revision_no,
                    t.file_data
                FROM month_ahead_file_uploads t
                INNER JOIN MaxRevisions mr
                    ON t.state_id = mr.state_id
                    AND t.from_date = mr.from_date
                    AND t.to_date = mr.to_date
                    AND t.revision_no = mr.max_revision
                WHERE
                    t.state_id IN ({2})
                ORDER BY t.from_date;
            '''.format(from_date, to_date, ','.join(map(str, expected_state_ids))))

            sr_month_data = cursor.fetchall()

            # Expand month-ahead forecast data by date range for each state
            expanded_month_forecast = []
            for record in sr_month_data:
                state_id, from_date, to_date, _, file_data = record
                if state_id not in expected_state_ids:
                    continue
                if file_data:
                    month_forecast_values = [forecast[3] for forecast in file_data]

                    current_date = from_date
                    while current_date <= to_date:
                        forecast_index = (current_date - from_date).days * 96
                        daily_forecast_96 = month_forecast_values[forecast_index:forecast_index + 96]
                        
                        if len(daily_forecast_96) == 96:
                            expanded_month_forecast.append({
                                "M_F_DATE": current_date.strftime("%Y-%m-%d"),
                                "MONTH_FORECAST": daily_forecast_96,
                                "STATE_ID": state_id
                            })
                        else:
                            print(f"Warning: Insufficient forecast data for {current_date.strftime('%Y-%m-%d')}")
                        
                        current_date += timedelta(days=1)

            # Convert expanded forecast data to DataFrame
            sr_expanded_month_forecast_df = pd.DataFrame(expanded_month_forecast)

            # Group by date and sum forecasts across states for each day
            sr_month_forecast_daily = (
                sr_expanded_month_forecast_df.groupby("M_F_DATE").apply(
                    lambda group: {
                        "MONTH_FORECAST": np.sum(np.vstack(group["MONTH_FORECAST"]), axis=0).tolist(),
                        "STATE_IDS": list(group["STATE_ID"])
                    }
                )
            ).apply(pd.Series).reset_index()

            formatted_month_forecast = []

            # Merge data frames
            date_df["D_F_DATE"] = pd.to_datetime(date_df["D_F_DATE"])
            sr_month_forecast_daily["M_F_DATE"] = pd.to_datetime(sr_month_forecast_daily["M_F_DATE"])
            sr_actual_day_df["A_DATE"] = pd.to_datetime(sr_actual_day_df["A_DATE"])

            sr_result_df = pd.merge(date_df, sr_month_forecast_daily, left_on="D_F_DATE", right_on="M_F_DATE", how="left")
            sr_result_df = pd.merge(sr_result_df, sr_actual_day_df, left_on="D_F_DATE", right_on="A_DATE", how="left")

            # Calculate MAPE with complete state data only
            sr_mape_dict = {"name": "Month Ahead", "data": []}
            for _, row in sr_result_df.iterrows():
                temp = {"x": row["A_DATE"].strftime('%Y-%m-%d')}
                forecast_state_ids = set(row["STATE_IDS_x"]) if isinstance(row["STATE_IDS_x"], list) else set()
                actual_state_ids = set(row["STATE_IDS_y"]) if isinstance(row["STATE_IDS_y"], list) else set()

                if forecast_state_ids != expected_state_ids or actual_state_ids != expected_state_ids:
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_month_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                elif (len(row["MONTH_FORECAST"]) != 96 or pd.isna(row["MONTH_FORECAST"]).any()):
                    temp["y"] = None
                    for i in range(96):
                        timestamp = row["D_F_DATE"] + timedelta(minutes=15 * i)
                        formatted_month_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": 0
                        })
                else:
                    day_actual_values = row['SCADA_DEMAND'] if all(x == 0 for x in row["DAY_ACTUAL"]) or (len(row["DAY_ACTUAL"]) != 96 or pd.isna(row["DAY_ACTUAL"]).any()) else row['DAY_ACTUAL']
                    mape_value = calculate_mape(day_actual_values, row["MONTH_FORECAST"])
                    temp["y"] = round(mape_value, 2) if not isinstance(mape_value, str) else None

                    base_date = row['D_F_DATE']
                    for i, forecast_value in enumerate(row["MONTH_FORECAST"]):
                        timestamp = base_date + timedelta(minutes=15 * i)
                        formatted_month_forecast.append({
                            "x": timestamp.strftime("%Y-%m-%d %H:%M"),
                            "y": round(forecast_value)
                        })

                sr_mape_dict["data"].append(temp)

            return formatted_month_forecast, sr_mape_dict


        day_forecast_data, actual_data, day_mape = day_ahead_mape(from_date, to_date, sr_actual_day_df)
        week_forecast, week_mape = week_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id)  # Replace with desired state ID or 6 for region
        month_forecast, month_mape = month_ahead_mape(from_date, to_date, sr_actual_day_df, selected_state_id)  # Replace with desired state ID or 6 for region
        intraday_forecast, intra_mape = intraday_mape(from_date, to_date, sr_actual_day_df)  # Replace with desired state ID or 6 for region

        # import pdb
        # pdb.set_trace()


        mape_list = []
        mape_list.append(day_mape)
        mape_list.append(week_mape)
        mape_list.append(month_mape)
        mape_list.append(intra_mape)


        forecast_list = []
        forecast_list.append({"name": "Day Ahead", "data": day_forecast_data})
        forecast_list.append({"name": "Week Ahead", "data": week_forecast})
        forecast_list.append({"name": "Month Ahead", "data": month_forecast})
        forecast_list.append({"name": "Actual", "data": actual_data})
        forecast_list.append({"name": "Intraday", "data": intraday_forecast})

        final_forecast_dict = {}
        final_forecast_dict['data'] = forecast_list 
        final_forecast_dict['title'] = f"Comparison between Actual, Intraday, Day Ahead, Week Ahead and Month Ahead data  from {from_date} to {to_date} for {state_name if state_name != 'ERLDC' else 'ER'}"
        
        # print(mape_list)

        # comp_data = {}  # Replace this with your `final_forecast_dict`

        # # Prepare the JSON object
        # json_content = {
        #     "status": "success",
        #     "data": mape_list,
        #     "title": "MAPE for the data between {0} and {1} for {2}".format(from_date, to_date, state_name if state_name != 'SRLDC' else 'SR'),
        #     "comp_data": final_forecast_dict
        # }

        # # File path to save the JSON file
        # file_path = "mape_data.json"  # Update the path as needed

        # # Write to the JSON file
        # with open(file_path, "w") as json_file:
        #     json.dump(json_content, json_file, indent=4)

        # print(f"JSON file saved at: {file_path}")


        

        return jsonify(status="success", data=mape_list, title="MAPE for the data between {0} and {1} for {2}".format(from_date, to_date, state_name if state_name != 'ERLDC' else 'ER' ), comp_data=final_forecast_dict)        

    except Exception as e:
        log_error("mapechart", e)
        import traceback
        print(traceback.format_exc())
        cursor.close()
        return jsonify(message="There is some problem in Fetching the Data! Please contact ERLDC IT", status="failure")







    



@app.route('/api/dayrangestatus', methods=['POST'])
@jwt_required()
@token_required
def dayRangeStatus():
    try:
        # Retrieve date range from request
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        params = request.get_json()
        print(params)

        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y')
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y')

        # SQL query to fetch data within the specified date range
        sql_query = """
                    WITH min_revision AS (
                    SELECT 
                        state_id, 
                        upload_date, 
                        MIN(revision_no) AS min_revision_no
                    FROM 
                        file_uploads
                    GROUP BY 
                        state_id, 
                        upload_date
                )
                SELECT 
                    states.state_name,
                    COALESCE(file_uploads.upload_date, %s) AS upload_date,
                    COALESCE(min_uploads.upload_time, NULL) AS upload_time,
                    CASE
                        WHEN min_uploads.upload_time IS NULL THEN 2  -- Not Uploaded
                        WHEN min_uploads.upload_time < (file_uploads.upload_date - INTERVAL '1 day' + INTERVAL '10 hours') THEN 1  -- Proper Upload before 10 AM on the previous day
                        ELSE 0  -- Late Upload
                    END AS upload_status_code,
                    COUNT(file_uploads.state_id) AS upload_count
                FROM 
                    states
                LEFT JOIN 
                    file_uploads ON states.state_id = file_uploads.state_id
                    AND file_uploads.upload_date BETWEEN %s AND %s
                LEFT JOIN (
                    SELECT 
                        file_uploads.state_id, 
                        file_uploads.upload_date, 
                        file_uploads.upload_time
                    FROM 
                        file_uploads
                    INNER JOIN min_revision ON 
                        file_uploads.state_id = min_revision.state_id
                        AND file_uploads.upload_date = min_revision.upload_date
                        AND file_uploads.revision_no = min_revision.min_revision_no
                ) AS min_uploads ON 
                    file_uploads.state_id = min_uploads.state_id 
                    AND file_uploads.upload_date = min_uploads.upload_date
                WHERE 
                    states.state_id IN (1,2,3,4,5,7)
                GROUP BY 
                    states.state_name, 
                    file_uploads.upload_date,
                    min_uploads.upload_time  
                ORDER BY
                    states.state_name,
                    upload_date DESC;
        """
        cursor.execute(sql_query, (end_date, start_date, end_date))
        results = cursor.fetchall()

        day_data = []
        date_range = [end_date - timedelta(days=i) for i in range((end_date - start_date).days + 1)]
        state_names = set(result[0] for result in results)

        for state_name in state_names:
            state_data = {"name": state_name, "data": []}
            for date in date_range:
                date_str = date.strftime('%Y-%m-%d')
                found_result = False
                for result in results:
                    if result[0] == state_name and result[1].strftime("%Y-%m-%d") == date_str:
                        upload_count = result[4]
                        upload_time = result[2].strftime("%Y-%m-%d %H:%M:%S") if result[2] is not None else None
                        upload_status_code = result[3]
                        found_result = True
                        break
                if not found_result:
                    upload_count = 0
                    upload_time = None
                    upload_status_code = 2  # Not Uploaded
                state_data["data"].append({
                    'x': date_str, 
                    'y': upload_status_code, 
                    'upload_time': upload_time, 
                    'upload_count': upload_count
                })
            day_data.append(state_data)

        day_dates = {
            "start_date": (start_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }

        return jsonify(day=day_data, status="success")

    except Exception as e:
        log_error("dayrangestatus", e)
        cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact ERLDC IT!", status="failure")





@app.route('/api/weekrangestatus', methods=['POST'])
@jwt_required()
@token_required
def weekRangeStatus():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Retrieve date range from request
        params = request.get_json()
        custom_start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y').date()
        custom_end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y').date()

        # Adjust start date to the start of the week (Monday)
        if custom_start_date.weekday() != 0:  # If the start date is not a Monday
            custom_start_date = custom_start_date - timedelta(days=custom_start_date.weekday())  # Shift to the previous Monday

        # Adjust the end date to include the entire last week if it falls mid-week
        if custom_end_date.weekday() != 6:  # If the end date is not a Sunday
            custom_end_date = custom_end_date + timedelta(days=(6 - custom_end_date.weekday()))  # Extend to the next Sunday

        # Generate the week ranges (from Monday to Sunday)
        week_range = []
        week_start = custom_start_date
        while week_start <= custom_end_date:
            week_range.append(week_start)
            week_start += timedelta(weeks=1)

        # SQL query for fetching data across weeks
        sql_query = """
            WITH min_revision AS (
                SELECT 
                    state_id, 
                    from_date, 
                    MIN(revision_no) AS min_revision_no
                FROM 
                    week_ahead_file_uploads
                WHERE
                    from_date BETWEEN %s AND %s
                GROUP BY 
                    state_id, 
                    from_date
            ), min_uploads AS (
                SELECT 
                    fu.state_id, 
                    fu.from_date, 
                    fu.upload_time,
                    fu.revision_no
                FROM 
                    week_ahead_file_uploads fu
                INNER JOIN min_revision mr ON
                    fu.state_id = mr.state_id
                    AND fu.from_date = mr.from_date
                    AND fu.revision_no = mr.min_revision_no
            )
            SELECT 
                states.state_name,
                mu.from_date AS week_start_date,
                mu.upload_time,
                CASE
                    WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                    WHEN mu.upload_time < DATE_TRUNC('week', mu.from_date - INTERVAL '1 week')  + INTERVAL '1 day' THEN 1  -- Uploaded before the Monday of the previous week
                    ELSE 0  -- Late Upload
                END AS upload_status_code
            FROM 
                states
            LEFT JOIN 
                min_uploads mu ON states.state_id = mu.state_id
            WHERE 
                states.state_id IN (1,2,3,4,5,7)
            GROUP BY 
                states.state_name, 
                mu.from_date,
                mu.upload_time  
            ORDER BY
                states.state_name,
                mu.from_date DESC;
        """
        
        cursor.execute(sql_query, (custom_start_date.strftime('%Y-%m-%d'), custom_end_date.strftime('%Y-%m-%d')))
        results = cursor.fetchall()

        # Fetch state names
        cursor.execute("SELECT state_name FROM states WHERE state_id IN (1,2,3,4,5,7)")
        all_states = [row[0] for row in cursor.fetchall()]

        # Initialize data structure for states
        week_data = [{"name": state_name, "data": []} for state_name in all_states]

        # Process each state and week range
        for state in week_data:
            for week_start in week_range:
                week_end = week_start + timedelta(days=6)  # End of the week
                # Filter the results for the current state and week
                upload_info = next((item for item in results if item[0] == state["name"] and item[1] is not None and week_start <= item[1] <= week_end), None)

                # Determine upload status for the week
                if upload_info:
                    upload_status_code = upload_info[3]
                    upload_time = upload_info[2].strftime('%Y-%m-%d %H:%M:%S') if upload_info[2] else None
                    upload_count = 1  # At least one upload exists
                else:
                    upload_status_code = 2  # Not Uploaded
                    upload_time = None
                    upload_count = 0

                # Append data for the week
                state["data"].append({
                    "x": f"{week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')}",
                    "y": upload_status_code,
                    "upload_time": upload_time,
                    "upload_count": upload_count
                })

            # Sort the state's data in descending order of week (latest first)
            state["data"] = sorted(state["data"], key=lambda x: x["x"], reverse=True)

        return jsonify(week=week_data, status="success")

    except Exception as e:
        log_error("weekrangestatus", e)
        cursor.close()
        return jsonify(message="There is a problem, please contact ERLDC IT!", status="failure")




@app.route('/api/monthrangestatus', methods=['POST'])
@jwt_required()
@token_required
def monthRangeStatus():
    try:
        # Get the current date
        today = datetime.now()

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        # Retrieve date range from request

        # Retrieve date range from request
        params = request.get_json()
        print(params)

        # Assuming 'params' is a dictionary that includes 'fromDate' and 'toDate'
        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y')
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y')

        # Calculate the start of the first month in the custom range
        start_of_first_month = start_date.replace(day=1)

        # Calculate the end of the last month in the custom range
        _, last_day = monthrange(end_date.year, end_date.month)
        end_of_last_month = end_date.replace(day=last_day)

        # Execute SQL Query
        sql_query = """
        WITH min_revision AS (
            SELECT 
                state_id, 
                DATE_TRUNC('month', from_date) AS from_month, 
                MIN(revision_no) AS min_revision_no
            FROM 
                month_ahead_file_uploads
            WHERE
                from_date BETWEEN %s AND %s
            GROUP BY 
                state_id, 
                from_month
        ), min_uploads AS (
            SELECT 
                fu.state_id, 
                DATE_TRUNC('month', fu.from_date) AS from_month, 
                fu.upload_time,
                fu.revision_no
            FROM 
                month_ahead_file_uploads fu
            INNER JOIN min_revision mr ON
                fu.state_id = mr.state_id
                AND DATE_TRUNC('month', fu.from_date) = mr.from_month
                AND fu.revision_no = mr.min_revision_no
        )
        SELECT 
            states.state_name,
            COALESCE(mu.from_month, %s) AS month_start_date,
            COALESCE(mu.upload_time, NULL) AS upload_time,
            CASE
                WHEN mu.upload_time IS NULL THEN 2  -- Not Uploaded
                WHEN mu.upload_time < DATE_TRUNC('month', mu.from_month) - INTERVAL '1 month' + INTERVAL '5 day' THEN 1  -- Uploaded on time
                ELSE 0  -- Late Upload
            END AS upload_status_code,
            COUNT(mu.state_id) AS upload_count
        FROM 
            states
        LEFT JOIN 
            min_uploads mu ON states.state_id = mu.state_id
        WHERE 
            states.state_id IN (1,2,3,4,5,7)
        GROUP BY 
            states.state_name, 
            mu.from_month,
            mu.upload_time  
        ORDER BY
            states.state_name,
            mu.from_month DESC;
        """
        cursor.execute(sql_query, (start_of_first_month.strftime('%Y-%m-%d'), end_of_last_month.strftime('%Y-%m-%d'), end_of_last_month.strftime('%Y-%m-%d')))
        results = cursor.fetchall()

        # Process the results
        month_range = [start_of_first_month + timedelta(days=32 * i) for i in range((end_of_last_month.year - start_of_first_month.year) * 12 + end_of_last_month.month - start_of_first_month.month + 1)]
        month_range = [date.replace(day=1) for date in month_range]
        all_states = set(result[0] for result in results)

        month_data = []
        # Process results and sort by latest month first
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for month_start in sorted(month_range, reverse=True):
                _, last_day_of_month = monthrange(month_start.year, month_start.month)
                month_end = month_start.replace(day=last_day_of_month)
                month_range_str = f"{month_start.strftime('%Y-%m-%d')} to {month_end.strftime('%Y-%m-%d')}"

                upload_count = 0
                upload_status_code = 2  # Default to Not Uploaded
                upload_time = None
                found_result = False

                # Check each result for the current state and month
                for result in results:
                    result_state, result_month_start, result_upload_time, result_status, result_count = result
                    if result_state == state_name and result_month_start.strftime('%Y-%m-%d') == month_start.strftime('%Y-%m-%d'):
                        upload_count = result_count
                        upload_time = result_upload_time.strftime("%Y-%m-%d %H:%M:%S") if result_upload_time else None
                        upload_status_code = result_status
                        found_result = True
                        break

                if not found_result:
                    upload_count = 0
                    upload_status_code = 2  # Not Uploaded

                state_data["data"].append({
                    'x': month_range_str,
                    'y': upload_status_code,
                    'upload_time': upload_time,
                    'upload_count': upload_count
                })

            month_data.append(state_data)

        # # Prepare JSON for frontend
        # month_dates = {
        #     "start_date": start_of_first_month.strftime('%Y-%m-%d'),
        #     "end_date": end_of_last_month.strftime('%Y-%m-%d')
        # }

        month_data
        return jsonify(month=month_data, status="success")

    except Exception as e:
        log_error("monthrangestatus", e)
        cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact ERLDC IT!")


@app.route('/api/yearrangestatus', methods=['POST'])
@jwt_required()
@token_required
def yearRangeStatus():
    try:
        # Get the current date
        today = datetime.now()

        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # Retrieve date range from request
        params = request.get_json()
        print(params)

        # Assuming 'params' is a dictionary that includes 'fromDate' and 'toDate'
        start_date = datetime.strptime(params["params"]["fromDate"], '%d/%m/%Y').date()
        end_date = datetime.strptime(params["params"]["toDate"], '%d/%m/%Y').date()

        # Loop to calculate the financial years intersecting with the given range
        financial_years = []
        current_year_start = datetime(start_date.year, 4, 1) if start_date.month >= 4 else datetime(start_date.year - 1, 4, 1)
        current_year_start = current_year_start.date()  # Ensure this is a date

        while current_year_start <= end_date:
            current_year_end = datetime(current_year_start.year + 1, 3, 31).date()  # Ensure this is also a date
            if current_year_end >= start_date:
                financial_years.append((current_year_start, current_year_end))
            current_year_start = datetime(current_year_start.year + 1, 4, 1).date()  # Next financial year start

        # SQL query to fetch relevant data
        sql_query = """
        SELECT 
            states.state_name,
            yafu.from_date,
            yafu.to_date,
            yafu.upload_time,
            CASE
                WHEN yafu.upload_time IS NULL THEN 2  -- Not Uploaded
                WHEN yafu.upload_time <= yafu.from_date + INTERVAL '5 months' - INTERVAL '1 day' THEN 1  -- Uploaded
                ELSE 0  -- Late Upload
            END AS upload_status,
            COUNT(yafu.state_id) AS upload_count
        FROM 
            states
        LEFT JOIN 
            year_ahead_file_uploads yafu ON states.state_id = yafu.state_id
        WHERE 
            states.state_id IN (1,2,3,4,5,7)
        GROUP BY 
            states.state_name, yafu.from_date, yafu.to_date, yafu.upload_time
        """
        cursor.execute(sql_query)
        results = cursor.fetchall()

        all_states = set(result[0] for result in results)

        # Process results
        year_data = []
        for state_name in all_states:
            state_data = {"name": state_name, "data": []}
            for fy_start, fy_end in financial_years:
                found = False
                for result in results:
                    result_state, from_date, to_date, upload_time, upload_status, upload_count = result
                    if result_state == state_name and from_date and to_date and from_date >= fy_start and to_date <= fy_end:
                        found = True
                        upload_time_formatted = upload_time.strftime("%Y-%m-%d %H:%M:%S") if upload_time else None
                        state_data["data"].append({
                            'x': f"{fy_start.strftime('%Y-%m-%d')} to {fy_end.strftime('%Y-%m-%d')}",
                            'y': upload_status,
                            'upload_time': upload_time_formatted,
                            'upload_count': upload_count
                        })
                if not found:
                    state_data["data"].append({
                        'x': f"{fy_start.strftime('%Y-%m-%d')} to {fy_end.strftime('%Y-%m-%d')}",
                        'y': 2,  # Not Uploaded
                        'upload_time': None,
                        'upload_count': 0
                    })
            year_data.append(state_data)

        # Prepare JSON data for the frontend
        date_range = {
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        }
        
        return jsonify(year=year_data, status="success")

    except Exception as e:
        log_error("yearrangestatus", e)
        cursor.close()
        # return jsonify(message="There is some problem in uploading the file! Please contact ERLDC IT", status="failure")
        return jsonify(message="There is a problem, please contact ERLDC IT!")










    
# Special /data endpoint only accessible by script token
@app.route('/api/data', methods=['POST'])
@jwt_required()
@token_required
def get_data():
    data = request.get_json()
    state_id = data.get('state')
    input_from_date = data.get('from_date')
    input_to_date = data.get('to_date')

    conn = psycopg2.connect(
    database="demand_forecast_states", user='ra_admin', 
    password='admin', host='localhost', port='5432'
    )

    cursor = conn.cursor()

    print(f"State ID: {state_id}, From Date: {input_from_date}, To Date: {input_to_date}")

    if not (state_id and input_from_date and input_to_date):
        return jsonify({'error': 'Missing data for state_id, from_date, or to_date'}), 400

    try:


        # Updated SQL query to filter by state_id, from_date, and to_date
        query = """
        SELECT
          state_id,
          from_date,
          to_date,
          revision_no,
          upload_time,
          file_data
        FROM (
          SELECT
            state_id,
            from_date,
            to_date,
            revision_no,
            upload_time,
            file_data,
            ROW_NUMBER() OVER (PARTITION BY state_id, from_date, to_date ORDER BY upload_time DESC) as rn
          FROM
            week_ahead_file_uploads
          WHERE state_id = %s AND from_date >= %s AND to_date <= %s
        ) sub
        WHERE sub.rn = 1;
        """

        cursor.execute(query, (state_id, input_from_date, input_to_date))
        records = cursor.fetchall()

        df_list = []

        results = []
        columns = ['Date', 'Block', 'Period', 'Forcasted Demand_MW (A)', 'From its own Sources Excl. Renewable_THERMAL_MW', 'From its own Sources Excl. Renewable_GAS_MW','From its own Sources Excl. Renewable_HYDRO_MW', 'From its own Sources Excl. Renewable_TOTAL (B)_MW', 'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 'From Renewable Sources_Other RES (biomass)_MW', 'From Renewable Sources_TOTAL (C)_MW', 'From ISGS & Other LTA & MTOA (D)_MW', 'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 'Total Availability  (F)= (B+C+D+E)_MW', 'Gap between Demand & Availability (G) = (A)-(F)  Surplus(-) / Deficit (+)_MW' , 'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 'Proposed Procurement_Through Power Exchange (I)_MW', 'Shortages after day ahead procurement from market (J) =(G)-(H+I)  Surplus(-) / Deficit (+)_MW', 'Relief through planned restrictions/ rostering/ power cuts (K)_MW','Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 'Reactive Power Forecast_MVar' ]

        for record in records:
            from_date = record[1]  # From Date
            to_date = record[2]    # To Date
            
            # Generate the Date column based on from_date and to_date
            date_range = generate_date_range(from_date, to_date)

            # Ensure the generated date list has the correct length to match the rows
            if len(date_range) != len(record[5]):
                raise ValueError(f"Generated date range length {len(date_range)} does not match file data length {len(record[5])}")

            # Replace "Invalid Date" with the actual date range
            file_data = record[5]
            for i, row in enumerate(file_data):
                row[0] = date_range[i]  # Assign the correct date to the first column
            
            result = {
                "State ID": record[0],
                "From Date": from_date.isoformat(),
                "To Date": to_date.isoformat(),
                "Revision No": record[3],
                "Upload Time": record[4].isoformat(),
                "File Data": file_data
            }

            # Create DataFrame from the updated File Data
            df_file_data = pd.DataFrame(result["File Data"], columns=columns)
            df_list.append(df_file_data)

        if not df_list:
            return jsonify({'error': 'No data present for this period'}), 500


        final_df = pd.concat(df_list, ignore_index=True)
        final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst = True)
        df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format

        cursor.close()
        conn.close()

        return jsonify(df_json)

    except Exception as e:
        print(f"Database connection or SQL execution error: {e}")
        log_error("data query", e)
        return jsonify({'error': 'Database query failed'}), 500



################ CONSOLIDATED DATA

@app.route("/api/fetchforecastdata", methods=['POST'])
@jwt_required()
@token_required
def get_forecast_data():
    try:
        data = request.get_json()
        state_id = data.get('state')
        input_from_date = data.get('from_date')
        input_to_date = data.get('to_date')
        data_type = data.get('data_type')

        conn = psycopg2.connect(    
        database="demand_forecast_states", user='ra_admin', 
        password='admin', host='localhost', port='5432'
        )

        cursor = conn.cursor()

        if not (state_id and input_from_date and input_to_date and data_type):
            return jsonify({'error': 'Missing data for state_id, from_date, or to_date or type'}), 400


        if data_type == 'day':
            # SQL query to filter by state_id and upload_date range
            query = """
                SELECT
                state_id,
                upload_date,
                revision_no,
                upload_time,
                file_data
                FROM (
                SELECT
                    state_id,
                    upload_date,
                    revision_no,
                    upload_time,
                    file_data,
                    ROW_NUMBER() OVER (PARTITION BY state_id, upload_date ORDER BY upload_time DESC) as rn
                FROM
                    file_uploads  -- Table without from_date and to_date, using upload_date
                WHERE state_id = %s AND upload_date BETWEEN %s AND %s
                ) sub
                WHERE sub.rn = 1;
            """

            # Execute the query with state_id, input_from_date, and input_to_date
            cursor.execute(query, (state_id, input_from_date, input_to_date))
            records = cursor.fetchall()

            df_list = []

            results = []
            columns = ['Date', 'Block', 'Period', 'Forecasted Demand_MW (A)', 
                    'From its own Sources Excl. Renewable_THERMAL_MW', 
                    'From its own Sources Excl. Renewable_GAS_MW',
                    'From its own Sources Excl. Renewable_HYDRO_MW', 
                    'From its own Sources Excl. Renewable_TOTAL (B)_MW', 
                    'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 
                    'From Renewable Sources_Other RES (biomass)_MW', 
                    'From Renewable Sources_TOTAL (C)_MW', 
                    'From ISGS & Other LTA & MTOA (D)_MW', 
                    'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 
                    'Total Availability  (F)= (B+C+D+E)_MW', 
                    'Gap between Demand & Availability (G) = (A)-(F) Surplus(-) / Deficit (+)_MW', 
                    'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 
                    'Proposed Procurement_Through Power Exchange (I)_MW', 
                    'Shortages after day ahead procurement from market (J) =(G)-(H+I) Surplus(-) / Deficit (+)_MW', 
                    'Relief through planned restrictions/ rostering/ power cuts (K)_MW',
                    'Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 
                    'Reactive Power Forecast_MVar']

            for record in records:
                upload_date = record[1]  # Upload Date

                # Generate a list of the same date to fill the 'Date' column for all 96 rows
                date_range = [upload_date] * 96  # Assuming 96 rows per date as per your data

                # Get the file_data from the record
                file_data = record[4]

                # Ensure the correct date is added to each row's first element
                for i, row in enumerate(file_data):
                    row.insert(0, date_range[i])  # Insert the date at the beginning of each row

                result = {
                    "State ID": record[0],
                    "Upload Date": upload_date.isoformat(),
                    "Revision No": record[2],
                    "Upload Time": record[3].isoformat(),
                    "File Data": file_data
                }

                # Now that date has been added to file_data, create the DataFrame
                df_file_data = pd.DataFrame(result["File Data"], columns=columns)
                df_list.append(df_file_data)

            if not df_list:
                return jsonify({'error': 'No data present for this period'}), 500

            # Concatenate all dataframes into one final dataframe
            final_df = pd.concat(df_list, ignore_index=True)
            
            # Convert 'Date' column to datetime format
            final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst=True)

            # Convert the final DataFrame to JSON
            df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format

            cursor.close()
            conn.close()


            
            return jsonify(df_json)

        elif data_type == 'week':
            # Updated SQL query to filter by state_id, from_date, and to_date
            query = """
            SELECT
            state_id,
            from_date,
            to_date,
            revision_no,
            upload_time,
            file_data
            FROM (
            SELECT
                state_id,
                from_date,
                to_date,
                revision_no,
                upload_time,
                file_data,
                ROW_NUMBER() OVER (PARTITION BY state_id, from_date, to_date ORDER BY upload_time DESC) as rn
            FROM
                week_ahead_file_uploads
            WHERE state_id = %s AND from_date >= %s AND to_date <= %s
            ) sub
            WHERE sub.rn = 1;
            """

            cursor.execute(query, (state_id, input_from_date, input_to_date))
            records = cursor.fetchall()

            df_list = []

            results = []
            columns = ['Date', 'Block', 'Period', 'Forcasted Demand_MW (A)', 'From its own Sources Excl. Renewable_THERMAL_MW', 'From its own Sources Excl. Renewable_GAS_MW','From its own Sources Excl. Renewable_HYDRO_MW', 'From its own Sources Excl. Renewable_TOTAL (B)_MW', 'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 'From Renewable Sources_Other RES (biomass)_MW', 'From Renewable Sources_TOTAL (C)_MW', 'From ISGS & Other LTA & MTOA (D)_MW', 'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 'Total Availability  (F)= (B+C+D+E)_MW', 'Gap between Demand & Availability (G) = (A)-(F)  Surplus(-) / Deficit (+)_MW' , 'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 'Proposed Procurement_Through Power Exchange (I)_MW', 'Shortages after day ahead procurement from market (J) =(G)-(H+I)  Surplus(-) / Deficit (+)_MW', 'Relief through planned restrictions/ rostering/ power cuts (K)_MW','Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 'Reactive Power Forecast_MVar' ]

            for record in records:
                from_date = record[1]  # From Date
                to_date = record[2]    # To Date
                
                # Generate the Date column based on from_date and to_date
                date_range = generate_date_range(from_date, to_date)

                # Ensure the generated date list has the correct length to match the rows
                if len(date_range) != len(record[5]):
                    raise ValueError(f"Generated date range length {len(date_range)} does not match file data length {len(record[5])}")

                # Replace "Invalid Date" with the actual date range
                file_data = record[5]
                for i, row in enumerate(file_data):
                    row[0] = date_range[i]  # Assign the correct date to the first column
                
                result = {
                    "State ID": record[0],
                    "From Date": from_date.isoformat(),
                    "To Date": to_date.isoformat(),
                    "Revision No": record[3],
                    "Upload Time": record[4].isoformat(),
                    "File Data": file_data
                }

                # Create DataFrame from the updated File Data
                df_file_data = pd.DataFrame(result["File Data"], columns=columns)
                df_list.append(df_file_data)

            if not df_list:
                return jsonify({'error': 'No data present for this period'}), 500


            final_df = pd.concat(df_list, ignore_index=True)
            final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst = True)
            df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format

            cursor.close()
            conn.close()


            return jsonify(df_json)

        elif data_type == 'month':
            # Updated SQL query to filter by state_id, from_date, and to_date
            query = """
            SELECT
            state_id,
            from_date,
            to_date,
            revision_no,
            upload_time,
            file_data
            FROM (
            SELECT
                state_id,
                from_date,
                to_date,
                revision_no,
                upload_time,
                file_data,
                ROW_NUMBER() OVER (PARTITION BY state_id, from_date, to_date ORDER BY upload_time DESC) as rn
            FROM
                month_ahead_file_uploads
            WHERE state_id = %s AND from_date >= %s AND to_date <= %s
            ) sub
            WHERE sub.rn = 1;
            """

            cursor.execute(query, (state_id, input_from_date, input_to_date))
            records = cursor.fetchall()

            df_list = []

            results = []
            columns = ['Date', 'Block', 'Period', 'Forcasted Demand_MW (A)', 'From its own Sources Excl. Renewable_THERMAL_MW', 'From its own Sources Excl. Renewable_GAS_MW','From its own Sources Excl. Renewable_HYDRO_MW', 'From its own Sources Excl. Renewable_TOTAL (B)_MW', 'From Renewable Sources_SOLAR_MW', 'From Renewable Sources_WIND_MW', 'From Renewable Sources_Other RES (biomass)_MW', 'From Renewable Sources_TOTAL (C)_MW', 'From ISGS & Other LTA & MTOA (D)_MW', 'From Bilateral Transaction (Advance+ FCFS) (E)_MW', 'Total Availability  (F)= (B+C+D+E)_MW', 'Gap between Demand & Availability (G) = (A)-(F)  Surplus(-) / Deficit (+)_MW' , 'Proposed Procurement_Under Bilateral Transaction (Day Ahead+ Contingency) (H)_MW', 'Proposed Procurement_Through Power Exchange (I)_MW', 'Shortages after day ahead procurement from market (J) =(G)-(H+I)  Surplus(-) / Deficit (+)_MW', 'Relief through planned restrictions/ rostering/ power cuts (K)_MW','Additional Load shedding proposed (L) = (J)-(K) Surplus(-) / Deficit (+)_MW', 'Reactive Power Forecast_MVar' ]

            for record in records:
                from_date = record[1]  # From Date
                to_date = record[2]    # To Date
                
                # Generate the Date column based on from_date and to_date
                date_range = generate_date_range(from_date, to_date)

                # Ensure the generated date list has the correct length to match the rows
                if len(date_range) != len(record[5]):
                    raise ValueError(f"Generated date range length {len(date_range)} does not match file data length {len(record[5])}")

                # Replace "Invalid Date" with the actual date range
                file_data = record[5]
                for i, row in enumerate(file_data):
                    row[0] = date_range[i]  # Assign the correct date to the first column
                
                result = {
                    "State ID": record[0],
                    "From Date": from_date.isoformat(),
                    "To Date": to_date.isoformat(),
                    "Revision No": record[3],
                    "Upload Time": record[4].isoformat(),
                    "File Data": file_data
                }

                # Create DataFrame from the updated File Data
                df_file_data = pd.DataFrame(result["File Data"], columns=columns)
                df_list.append(df_file_data)

            if not df_list:
                return jsonify({'error': 'No data present for this period'}), 500


            final_df = pd.concat(df_list, ignore_index=True)
            final_df['Date'] = pd.to_datetime(final_df['Date'], dayfirst = True)
            df_json = final_df.to_json(orient='records', date_format='iso')  # 'records' is good for row-wise format
            
            return jsonify(df_json)

    except Exception as e:
        print(f"Some Problem with the API: {e}")
        log_error("forecast data fetch", e)
        return jsonify({'error': 'Database query failed'}), 500




    




##################### LINEFLOWS REPORT



@app.route("/api/reports/lineflows", methods=["POST"])
@jwt_required()
@token_required
def displayLineFlows():
    try:
        conn2 = psycopg2.connect(
        database="MDPNew", user = 'prasadbabu',
        password = 'BabuPrasad#123', host = '10.0.100.219', port = '5432' 
        )
        cur = conn2.cursor()
        input_dat = request.get_json()
        start_date = datetime.strptime(input_dat['params']['fromDate'].replace('"', ''), '%d/%m/%Y').date()
        end_date = datetime.strptime(input_dat['params']['toDate'].replace('"', ''), '%d/%m/%Y').date()


        # Get columns data
        cur.execute("""
            SELECT col_name, out_mwh_columns.formula as formula, col_seq_no, item, col_type, out_mwh_columns.startdatetime as startdatetime, out_mwh_columns.enddatetime as enddatetime
            FROM out_mwh_config join out_mwh_columns on out_mwh_config.id = out_mwh_columns.item_fk_id
            WHERE item = 'LFL_MWH' AND out_mwh_columns.startdatetime <= %s;
        """, (end_date,))
        columns_data = cur.fetchall()
        columns_df = pd.DataFrame(columns_data, columns=['col_name', 'formula', 'col_seq_no', 'item', 'col_type', 'startdatetime', 'enddatetime'])

        # Regular expression pattern
        pattern = r"[-+]?[A-Za-z0-9]+-[A-Za-z0-9]+\s*"
        for i, col in columns_df.iterrows():
            if col['formula']:
                matches = re.findall(pattern, col['formula'])
                if col['col_type'] in ['SINGLE', 'FREQ'] and len(matches) >= 1:
                    columns_df.at[i, 'neg_factor'] = -1 if col['formula'][0] == '-' else 1
                    columns_df.at[i, 'formula'] = matches[0]
                else:
                    columns_df.at[i, 'neg_factor'] = 1

        columns_df = columns_df.sort_values('col_seq_no').copy()

        # print(columns_df)

        # Get fict computed list
        cur.execute("""
            SELECT DISTINCT short_location, startdate, time, mwh
            FROM fict_computation
            WHERE startdate BETWEEN %s AND %s;
        """, (start_date, end_date))
        fict_computed_list = pd.DataFrame(cur.fetchall(), columns=['short_location', 'startdate', 'time', 'mwh'])

        # print(fict_computed_list.head())

        # Get NPC meter data
        cur.execute("""
            SELECT DISTINCT startdate, time, meterno, mwh
            FROM npcmeterdata
            WHERE startdate BETWEEN %s AND %s;
        """, (start_date, end_date))
        queryset_npc = pd.DataFrame(cur.fetchall(), columns=['startdate', 'time', 'meterno', 'mwh'])

        # Get replaced meter data
        cur.execute("""
            SELECT DISTINCT find_meterno, startdate, time, mwh
            FROM replaced_meter_data
            WHERE startdate BETWEEN %s AND %s;
        """, (start_date, end_date))
        replaced_query_set = pd.DataFrame(cur.fetchall(), columns=['find_meterno', 'startdate', 'time', 'mwh'])

        # Get fict meters locations
        cur.execute("""
            SELECT short_location
            FROM create_fictmeter
            WHERE end_date_time >= CURRENT_TIMESTAMP OR end_date_time IS NULL;
        """)
        fict_meters_locations = [loc[0] for loc in cur.fetchall()]

        # Get real meter data
        cur.execute("""
            SELECT meter_number, short_location, start_date_time, end_date_time
            FROM create_meter;
        """)
        realmeter_obj = pd.DataFrame(cur.fetchall(), columns=['meter_number', 'short_location', 'start_date_time', 'end_date_time'])        

        # print(realmeter_obj.head())

            # folder creation
        # directory=create_folder(start_date,end_date)
        entity_df=generate_dataframe_forrange(start_date,end_date)

        for _,col in columns_df.iterrows():
            # below is to check whether column is Single fict like (AP-NEW)
            if col['col_type'] in ['SINGLE'] :
                # this column is there but no fict formula so just keep the column in output csv
                if col['formula']!= None and col['formula']!='' and len(col['formula']) > 0 :
                    # checking fict meter first
                    full_data_df = fict_computed_list[(fict_computed_list["short_location"] == col["formula"]) & (fict_computed_list["startdate"].between(start_date, end_date, inclusive="both"))][["startdate", "time", "mwh"]]
                    full_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': 'mwh1'}, inplace=True)

                    if not full_data_df.empty:
                        #replace None with --
                        full_data_df.fillna('--',inplace=True)
                        entity_df=pd.merge(entity_df, full_data_df, on=['DATE', 'TIME'], how='left')
                        entity_df.rename(columns={'mwh1':col['col_name']} , inplace=True)
                    elif col['formula'] in fict_meters_locations:
                        # fict meter but not computed
                        entity_df=_column(entity_df , col['col_name'] )
                    else:
                        # if it is real meter but replaced for particular week
                        # replaced_meter_data_df=pd.DataFrame(replaced_queryset.filter(find_meterno=col['outmwhcolumns__formula']).values(**{'DATE':F('startdate'),'TIME':F('time'),'mwh1':F('mwh') }),columns=['DATE','TIME','mwh1'] )
                        replaced_meter_data_df = replaced_query_set[(replaced_query_set["find_meterno"] == col["formula"]) & (replaced_query_set["startdate"].between(start_date, end_date, inclusive="both"))][["startdate", "time", "mwh"]]
                        replaced_meter_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': 'mwh1'}, inplace=True)

                        # first get the meterno corresponding short location
                        # meterno_obj=realmeter_obj.filter( Q(short_location=col['outmwhcolumns__formula']) ,( Q(end_date_time__range=[start_date,end_date] )| Q(end_date_time__isnull=True) ) )
                        # realmeter_obj['end_date_time'] = realmeter_obj['end_date_time'].dt.date
                        
                        meterno_obj = realmeter_obj[(realmeter_obj["short_location"] == col['formula']) & ((realmeter_obj["end_date_time"].between(pd.to_datetime(start_date).tz_localize('UTC'), pd.to_datetime(end_date).tz_localize('UTC'), inclusive="both")) | (realmeter_obj["end_date_time"].isnull()) )]

                        real_meter_data_df=pd.DataFrame([],columns=['DATE','TIME'])


                        for m, mtr in meterno_obj.iterrows():
                            # temp_multi_real_df=pd.DataFrame( queryset_npc.filter(meterno=mtr.meter_number).values(**{'DATE':F('startdate'),'TIME':F('time'),'mwh1':F('mwh') }) ,columns=['DATE','TIME','mwh1'])
                            
                            temp_multi_real_df = queryset_npc[queryset_npc["meterno"] == mtr["meter_number"]][["startdate", "time", "mwh"]]
                            temp_multi_real_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': 'mwh1'}, inplace=True)

                            temp_multi_real_df=check_start_enddate_col(temp_multi_real_df ,mtr["start_date_time"] , mtr["end_date_time"],'mwh1')

                            real_meter_data_df=pd.concat([real_meter_data_df,temp_multi_real_df])
                        
                        replaced_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)
                        real_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)

                        if not real_meter_data_df.empty:
                            # Update values in real_meter_data_df with corresponding values from replaced_meter_data_df based on the index (DATE, TIME)
                            real_meter_data_df.update(replaced_meter_data_df)
                            # Reset the index to columns
                            real_meter_data_df.reset_index(inplace=True)
                            
                            # print("ENTITY_DF")
                            # print(entity_df.head())
                            # print("REAL_METER_DF")
                            # print(real_meter_data_df.head())

                            entity_df=pd.merge(entity_df, real_meter_data_df[["DATE", "TIME", "mwh1"]], on=['DATE', 'TIME'], how='left')
                            entity_df.rename(columns={'mwh1':col['col_name']} , inplace=True)
                        elif not replaced_meter_data_df.empty:
                            # if real meter data not present but replaced meter data presents then updat entity_df with only replaced duration
                            # Reset the index to columns
                            replaced_meter_data_df.reset_index(inplace=True)

                            # print("ENTITY_DF")
                            # print(entity_df.head())
                            # print("REPLACED_METER_DF")
                            # print(replaced_meter_df.head())

                            entity_df=pd.merge(entity_df, replaced_meter_data_df, on=['DATE', 'TIME'], how='left')
                            entity_df.rename(columns={'mwh1':col['col_name']} , inplace=True)
                        else:
                            entity_df=_column(entity_df , col['col_name'] )
                    
                    # finally multiplying with negative factor
                    entity_df[col['col_name']]=(entity_df[col['col_name']]) * int(col['neg_factor']) 

                else: 
                    # this else no fict formula but to keep the column in output file to match the srpc format
                    entity_df=_column(entity_df , col['col_name'] )

            elif col['col_type'] in ['MULTI'] :
                short_locations=set(split_string(col['formula']))
                multi_df=generate_dataframe_forrange(start_date,end_date)


                for loct in short_locations:
                    real_meter_data_df=generate_dataframe_forrange(start_date,end_date)
                    real_meter_data_df[loct]=None

                    modified_loc=loct.replace('(','').replace(')','')
                    # fist check if it is fict meter else real meter
                    # multi_fict_data_df=pd.DataFrame( fict_computed_list.filter(short_location= modified_loc , startdate__range=[start_date,end_date]).values(**{'DATE':F('startdate'),'TIME':F('time'), loct:F('mwh') } ) , columns=['DATE','TIME',loct])
                    multi_fict_data_df = fict_computed_list[(fict_computed_list["short_location"] == modified_loc) & (fict_computed_list["startdate"].between(start_date, end_date, inclusive="both"))]
                    multi_fict_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': loct}, inplace=True)

                    if not multi_fict_data_df.empty:
                        multi_fict_data_df.fillna('--',inplace=True)
                        multi_df=pd.merge(multi_df, multi_fict_data_df, on=['DATE', 'TIME'], how='inner')

                    elif modified_loc in fict_meters_locations:
                        multi_df=_column(multi_df , loct )

                    else:
                        # if it is real meter but replaced for particular week
                        # replaced_meter_data_df=pd.DataFrame(replaced_queryset.filter(find_meterno=modified_loc).values(**{'DATE':F('startdate'),'TIME':F('time'),loct:F('mwh') }),columns=['DATE','TIME',loct] )

                        replaced_meter_data_df = replaced_query_set[(replaced_query_set["find_meterno"] == modified_loc) ][["startdate", "time", "mwh"]]
                        replaced_meter_data_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': loct}, inplace=True)

                        # meterno_obj=realmeter_obj.filter( Q(short_location=modified_loc) , ( Q(end_date_time__range=[start_date,end_date] )| Q(end_date_time__isnull=True) ) )

                        meterno_obj = realmeter_obj[(realmeter_obj["short_location"] == modified_loc) & ((realmeter_obj["end_date_time"].between(pd.to_datetime(start_date).tz_localize('UTC'), pd.to_datetime(end_date).tz_localize('UTC'), inclusive="both")) | (realmeter_obj["end_date_time"].isnull()) )]


                        temp_multi_df1=pd.DataFrame([],columns=['DATE','TIME'])

                        for m, mtr in meterno_obj.iterrows():
                            # temp_multi_real_df=pd.DataFrame( queryset_npc.filter(meterno=mtr.meter_number).values(**{'DATE':F('startdate'),'TIME':F('time'),'mwh1':F('mwh') }) ,columns=['DATE','TIME','mwh1'])
                            
                            temp_multi_real_df = queryset_npc[queryset_npc["meterno"] == mtr["meter_number"]][["startdate", "time", "mwh"]]
                            temp_multi_real_df.rename(columns={'startdate': 'DATE', 'time': 'TIME', 'mwh': loct}, inplace=True)

                            temp_multi_real_df=check_start_enddate_col(temp_multi_real_df ,mtr["start_date_time"] , mtr["end_date_time"],loct)

                            temp_multi_df1=pd.concat([temp_multi_df1,temp_multi_real_df])
                        
                        real_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)
                        temp_multi_df1.set_index(['DATE', 'TIME'], inplace=True)
                        # Must set_index before using update method
                        real_meter_data_df.update(temp_multi_df1)
                        
                        replaced_meter_data_df.set_index(['DATE', 'TIME'], inplace=True)
                        

                        if not real_meter_data_df.empty:
                            # Update values in real_meter_data_df with corresponding values from replaced_meter_data_df based on the index (DATE, TIME)
                            real_meter_data_df.update(replaced_meter_data_df)
                            # Reset the index to columns
                            real_meter_data_df.reset_index(inplace=True)

                            multi_df=pd.merge(multi_df, real_meter_data_df[["DATE", "TIME", loct]], on=['DATE', 'TIME'], how='left')
                        elif not replaced_meter_data_df.empty:
                            # if real meter data not present but replaced meter data presents then updat entity_df with only replaced duration
                            # Reset the index to columns
                            replaced_meter_data_df.reset_index(inplace=True)

                            multi_df=pd.merge(multi_df, replaced_meter_data_df, on=['DATE', 'TIME'], how='left')   
                        else:  
                            multi_df=_column(multi_df , loct )

                if not multi_df.empty:
                    multi_df[col['col_name']] = multi_df.apply(calculate_formula,args=(short_locations, col['formula']), axis=1)
                else: 
                    multi_df=_column(multi_df , col['col_name'] )

                multi_df.drop(columns=short_locations,inplace=True)
                entity_df=pd.merge(entity_df, multi_df, on=['DATE', 'TIME'], how='left')


            else:
                pass
            
            #skip if col_type is DATE,TIME and TOTAL
            if col['col_type'] not in ['DATE','TIME','TOTAL']:
                # if startdatetime is greater than current date then putting -- as mwh1 else original data 
                new_frame_df=entity_df.copy()
                new_frame_df['DATETIME'] = pd.to_datetime(new_frame_df['DATE'].astype(str) + ' ' + new_frame_df['TIME'].astype(str) ,  errors='coerce')
                new_frame_df['DATETIME'] = new_frame_df['DATETIME'].apply(lambda x: x.tz_localize(None))

                temp_timestamp=pd.to_datetime(col['startdatetime']).tz_convert(None)
                temp_endtimestamp=pd.to_datetime(col['enddatetime'],errors='coerce')
                
                # temp_timestamp = temp_timestamp.apply(lambda x: x.tz_localize(None))
                # temp_endtimestamp = temp_endtimestamp.apply(lambda x: x.tz_localize(None))
            
                # print(new_frame_df.dtypes)

                if pd.isnull(temp_endtimestamp):
                    # Handle the invalid timestamps by replacing them with a default value
                    default_endtimestamp = pd.to_datetime('2050-01-01')
                else:
                    default_endtimestamp = temp_endtimestamp.tz_convert(None)

                record_dt_time_condition=new_frame_df['DATETIME'].apply(lambda dt: dt >= temp_timestamp and dt <= default_endtimestamp )

                new_frame_df.loc[~(record_dt_time_condition), col['col_name']] = '--'
                new_frame_df.drop(columns=['DATETIME'],inplace=True)
                entity_df=new_frame_df.copy()

        entity_df=entity_df.fillna('--')
        new_frame_df=entity_df.copy()

        if not new_frame_df.empty:
            new_frame_df.set_index(['DATE', 'TIME'], inplace=True)
            # Specify columns to include in calculations
            columns_to_include = new_frame_df.columns.difference(['DATE', 'TIME'])

            new_frame_df[new_frame_df.columns[2:]] = new_frame_df[new_frame_df.columns[2:]].apply(pd.to_numeric, errors='coerce')

            # Apply the function column-wise on the selected columns
            result_df = new_frame_df[columns_to_include].apply(calculate_column_stats, axis=0)
            try:
                # to rearrange the column order
                result_df = result_df[entity_df.columns[2:]]
            except Exception as e: 
                extractdb_errormsg(e)
            # Transpose the result DataFrame
            result_df = result_df.T
            result_df.reset_index(inplace=True)
            result_df.rename( columns={'Pos_Sum':'Export (MU)','Pos_Max':'Maximum Power Flow (MW)','Neg_Sum':'Import (MU)','Neg_Max':'Maximum Power Flow (MW)(I)' ,'index':'NAME OF THE LINE'},inplace=True ) 
        else:
            result_df=pd.DataFrame(['No data found , Please check'])

        filename='line_flows_'+start_date.strftime('%d-%m-%y')+'.xlsx'
        
        entity_df.set_index('DATE', inplace=True)

        df = entity_df
        df.columns = [col[:-2] if col.endswith('.1') else col for col in df.columns]
        df.drop(columns=['TIME'], inplace=True)
        df = df.applymap(lambda x: 0 if isinstance(x, str) else x)
        def calculate_positive_negative_sum(df, column_name):
            positive_sum = df[df[column_name] > 0][column_name].groupby(level=0).sum()
            negative_sum = df[df[column_name] < 0][column_name].groupby(level=0).sum()
            return positive_sum, negative_sum
        df2=pd.DataFrame()
        positive_sums = {}
        negative_sums = {}
        all_dates = df.index.unique()

        for column_name in df.columns:
            positive_sum, negative_sum = calculate_positive_negative_sum(df, column_name)
            positive_sum = positive_sum.reindex(all_dates, fill_value=0)
            negative_sum = negative_sum.reindex(all_dates, fill_value=0)
            df2[column_name + ' imp'] = positive_sum
            df2[column_name + ' exp'] = negative_sum
        df2= df2.div(1000)
        interreg=["RGDM-CHNDPR imp","RGDM-CHNDPR exp","765KV RCR-SLPR I&II imp","765KV RCR-SLPR I&II exp","HVDC GAJUWAKA imp","HVDC GAJUWAKA exp","HVDC TAL-KLR imp","HVDC TAL-KLR exp","765KV KDG-KLPR I&II imp","765KV KDG-KLPR I&II exp","ANG-SKLM imp","ANG-SKLM exp","WAR-NZB imp","WAR-NZB exp","RGH-PUG imp","RGH-PUG exp","220KV AMBWDI-XLDM imp","220KV AMBWDI-XLDM exp","220KV AMBWDI-PONDA imp","220KV AMBWDI-PONDA exp","765KV WAR-WRNGL I&II.1 imp","765KV WAR-WRNGL I&II.1 exp"]
        column_map ={"BHADRAVATHI IMP":"RGDM-CHNDPR imp","BHADRAVATHI EXP":"RGDM-CHNDPR exp","RCR-SOLPR IMP":"765KV RCR-SLPR I&II imp","RCR-SLPR EXP":"765KV RCR-SLPR I&II exp","GAZUWAKA IMP":"HVDC GAJUWAKA imp","GAZUWAKA EXP":"HVDC GAJUWAKA exp","TAL-KOLAR":"HVDC TAL-KLR imp","TAL-KOLAR EXP":"HVDC TAL-KLR exp","KUDGI-KOLAPUR IMP":"765KV KDG-KLPR I&II imp","KUDGI-KOLAPUR EXP":"765KV KDG-KLPR I&II exp","Angul-Srikakulam IMP":"ANG-SKLM imp","Angul-Srikakulam EXP":"ANG-SKLM exp","765kV Wardha-Nizamabad 1 & 2 IMP":"WAR-NZB imp","765kV Wardha-Nizamabad 1 & 2 EXP":"WAR-NZB exp","Raighar-Pugalur HVDC IMP":"RGH-PUG imp","Raighar-Pugalur HVDC EXP":"RGH-PUG exp","Ambewadi-Xeldom IMP":"220KV AMBWDI-XLDM imp","Ambewadi-Xeldom EXP":"220KV AMBWDI-XLDM exp","Ambewadi-Ponda IMP":"220KV AMBWDI-PONDA imp","Ambewadi-Ponda EXP":"220KV AMBWDI-PONDA exp","765KV WAR-WRNGL I&II IMP":"765KV WAR-WRNGL I&II imp","765KV WAR-WRNGL I&II EXP":"765KV WAR-WRNGL I&II exp"}

        # Load the workbook and select the desired sheet
        wb = openpyxl.load_workbook(os.path.join(shared_drive_path,"Lineflows Report work","LineFlows.xlsx"))
        # os.path.join(shared_drive_path
        sheet = wb['SUMMARY']

        # Clear the contents of the range A2:W32
        for row in sheet.iter_rows(min_row=2, max_row=32, min_col=1, max_col=23):  # A2:W32 is 2nd row to 32nd, and columns A to W (23 columns)
            for cell in row:
                cell.value = None

        # Assuming df2 is your DataFrame, converting the index to datetime
        index_datetime = pd.to_datetime(df2.index)

        # Writing dates to column A starting from row 2
        for row_num, date in enumerate(index_datetime.strftime('%d-%m-%Y'), start=2):
            sheet.cell(row=row_num, column=1, value=date)

        # Prepare JSON output
        json_output = []

        # Assuming column_map is a dictionary that maps your output columns to DataFrame columns
        for output_column, df_column in column_map.items():
            try:
                # Find the column index for the output column in the first row
                output_column_index = None
                for col_num, cell in enumerate(sheet[1], start=1):  # Sheet row 1
                    if cell.value == output_column:
                        output_column_index = col_num
                        break
                
                if output_column_index is None:
                    continue  # Skip if the column wasn't found
                
                # Writing data from the DataFrame to the corresponding column
                data = df2[df_column].values.tolist()
                for row_num, value in enumerate(data, start=2):
                    sheet.cell(row=row_num, column=output_column_index, value=value if pd.notna(value) else None)

                # Prepare the JSON output data
                data_entries = [{'x': date.strftime('%Y-%m-%d'), 'y': round(value,4) if pd.notna(value) else ''} for date, value in zip(df2.index, df2[df_column])]
                json_output.append({
                    'name': df_column,
                    'data': data_entries
                })
            except Exception as e:
                print(e)
                continue

        # Save the workbook with a timestamped name
        file_name = os.path.join(shared_drive_path,"Lineflows Report work","lineflows_report_{0}.xlsx".format(datetime.now().strftime('%Y%m%d_%H%M%S')))
        
        wb.save(file_name)
        wb.close()

        # Return the response in the same format as the original
        return jsonify({
            'status': 'success',
            "message": "Report generated successfully!",
            "file_link": file_name,
            "data": json_output,
            "title": "Lineflows for the period between {0} and {1}".format(input_dat['params']['fromDate'], input_dat['params']['toDate'])
        }), 200



    except Exception as e:
        log_error("displaylineflows", e)
        cursor.close()
        return jsonify(message="There is some problem in Fetching the data! Please contact ERLDC IT", status="failure")
        # return jsonify({"status": "failure", "error": str(e), "file": fname, "line": exc_tb.tb_lineno}), 500




        

@app.route("/api/reports/downloadlineflows", methods=["POST"])
@jwt_required()
@token_required
def downloadLineFlows():
    try:
        # Generate a unique filename with a timestamp
        download_link_json = request.get_json()
        download_link = download_link_json['downloadLink']

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'lineflows_report_{timestamp}.xlsx'
        file_path = download_link

        # Example logic for saving the Excel file
        # wb.save(file_path)
        # wb.close()

        # Directly send the file in the response
        return send_file(file_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        log_error("downloadlineflows", e)
        cursor.close()
        return jsonify(message="There is some problem in downloading the file! Please contact ERLDC IT", status="failure")
        # return jsonify({'status': 'failure', 'error': str(e)}), 500


@app.route("/api/reports/fetchmdpdescription", methods=["GET"])
@jwt_required()
@token_required
def fetchMdpDescription():
    try:
        # Establish database connection
        conn = psycopg2.connect(
            **mdp_db_params
        )
        cur = conn.cursor()

        # Execute the query to fetch distinct descriptions
        query = "SELECT DISTINCT description FROM create_fictmeter;"
        cur.execute(query)

        # Fetch all results
        result = cur.fetchall()

        # Close the cursor and connection
        cur.close()
        conn.close()

        # Format the results as a list of descriptions
        descriptions = [row[0] for row in result]


        return jsonify(data=descriptions, message="Successfully fetched descriptions!", status="success")

    except Exception as e:
        # Log the error and return a failure message
        log_error("fetchmdpdescription", e)
        return jsonify(message="There is some problem in fetching the names! Please contact ERLDC IT", status="failure")




@app.route("/api/reports/mdpdescriptiondata", methods=["POST"])
@jwt_required()
@token_required
def MdpDescriptionData():
    try:
        # Parse request data
        data = request.get_json()['params']
        element_list = data.get('elementList')  # List of elements
        input_from_date = data.get('fromDate')  # From date in DD/MM/YYYY format
        input_to_date = data.get('toDate')  # To date in DD/MM/YYYY format

        # Convert input dates to YYYY-MM-DD format
        from_date = datetime.strptime(input_from_date, "%d/%m/%Y").strftime("%Y-%m-%d")
        to_date = datetime.strptime(input_to_date, "%d/%m/%Y").strftime("%Y-%m-%d")

        # Establish database connection
        conn = psycopg2.connect(**mdp_db_params)
        cur = conn.cursor()

        # Prepare results for each element
        results = []
        for element in element_list:
            query = """
                SELECT
                    TO_CHAR(fict_computation."startdate" + fict_computation."time"::interval, 'YYYY-MM-DD HH24:MI') AS timestamp,
                    fict_computation."mw"
                FROM
                    public.fict_computation
                WHERE
                    fict_computation."short_location" = (
                        SELECT DISTINCT short_location FROM public.create_fictmeter WHERE description = %s
                    )
                    AND fict_computation."startdate" BETWEEN %s AND %s
                ORDER BY fict_computation."startdate", fict_computation."time"
            """
            cur.execute(query, (element, from_date, to_date))
            data_records = cur.fetchall()

            # Format the data
            formatted_data = [{'x': record[0], 'y': round(record[1], 2)} for record in data_records]
            results.append({'name': element, 'data': formatted_data})

        # Close the cursor and connection
        cur.close()
        conn.close()

        # Create a comma-separated list of elements for the title
        comma_separated_list = ', '.join(element_list)

        # Return the formatted results
        return jsonify(
            results=results,
            message="API Fetch Successful!",
            title="Showing Data for {} for the period between {} and {}".format(
                comma_separated_list, input_from_date, input_to_date
            ),
            status="success"
        )

    except Exception as e:
        # Log the error and return a failure message
        print(f"Error in mdpdescriptiondata: {e}")
        return jsonify(
            message="There is some problem in fetching the data! Please contact ERLDC IT",
            status="failure"
        )



if __name__ == '__main__':
    cors = CORS(app)
    app.run(host='0.0.0.0', port=5000, debug=False)






